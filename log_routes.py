"""
Журнал обращений — аналог Google Sheets.
Данные берутся из dialogs + dialog_analysis + raw_dialogs.
Редактируемые поля (Организатор, Ответственный, Результат, Комментарий)
хранятся в support_log_edits (ReplacingMergeTree).

Маршруты:
  GET  /log                  — HTML-страница
  GET  /api/log              — JSON: список строк + список операторов
  POST /api/log/{chat_id}    — сохранить правки
"""

import io
import json
import os
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from pydantic import BaseModel
from auth import require_user

router = APIRouter()

CH_HOST     = os.getenv("CH_HOST", "localhost")
CH_PORT     = int(os.getenv("CH_PORT", "8123"))
CH_USER     = os.getenv("CH_USER", "default")
CH_PASSWORD = os.getenv("CH_PASSWORD", "")
CH_DATABASE = os.getenv("CH_DATABASE", "default")


def _params():
    return urllib.parse.urlencode({
        "user": CH_USER, "password": CH_PASSWORD, "database": CH_DATABASE,
    })


def ch_query(sql: str) -> list:
    url = f"http://{CH_HOST}:{CH_PORT}/?{_params()}"
    req = urllib.request.Request(url, data=sql.encode("utf-8"), method="POST")
    try:
        resp = urllib.request.urlopen(req, timeout=15)
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"CH {e.code}: {e.read().decode()[:300]}")
    rows = []
    for line in resp.read().decode().strip().splitlines():
        if line:
            try:
                rows.append(json.loads(line))
            except Exception:
                pass
    return rows


def ch_execute(sql: str, data: bytes = None):
    if data is not None:
        # INSERT: SQL в URL query-параметре, данные в теле
        params = urllib.parse.urlencode({
            "query": sql,
            "user": CH_USER, "password": CH_PASSWORD, "database": CH_DATABASE,
        })
        url  = f"http://{CH_HOST}:{CH_PORT}/?{params}"
        body = data
    else:
        url  = f"http://{CH_HOST}:{CH_PORT}/?{_params()}"
        body = sql.encode("utf-8")
    req = urllib.request.Request(url, data=body, method="POST")
    try:
        urllib.request.urlopen(req, timeout=15)
    except urllib.error.HTTPError as e:
        raise RuntimeError(f"CH {e.code}: {e.read().decode()[:300]}")


def ensure_table():
    """Создаёт вспомогательные таблицы если их нет, добавляет новые колонки."""
    ch_execute("""
        CREATE TABLE IF NOT EXISTS support_log_edits (
            chat_id     UInt64,
            updated_at  DateTime DEFAULT now(),
            organizer   String   DEFAULT '',
            responsible String   DEFAULT '',
            result      String   DEFAULT '',
            comment     String   DEFAULT ''
        ) ENGINE = ReplacingMergeTree(updated_at)
        ORDER BY chat_id
    """)
    for col in ["source_type", "category", "subcategory", "comment_manager",
                "comment_author", "comment_manager_author"]:
        try:
            ch_execute(f"ALTER TABLE support_log_edits ADD COLUMN IF NOT EXISTS {col} String DEFAULT ''")
        except Exception:
            pass
    for col in ["rating", "complexity"]:
        try:
            ch_execute(f"ALTER TABLE support_log_edits ADD COLUMN IF NOT EXISTS {col} UInt8 DEFAULT 0")
        except Exception:
            pass
    ch_execute("""
        CREATE TABLE IF NOT EXISTS day_tracker_edits (
            chat_id          UInt64,
            updated_at       DateTime DEFAULT now(),
            responsible_dept String   DEFAULT ''
        ) ENGINE = ReplacingMergeTree(updated_at)
        ORDER BY chat_id
    """)
    ensure_manual_table()
    ensure_col_perms_table()


def ensure_manual_table():
    """Временная таблица для ручных строк журнала (до интеграции новых каналов)."""
    ch_execute("""
        CREATE TABLE IF NOT EXISTS manual_log_entries (
            id              UInt64,
            updated_at      DateTime DEFAULT now(),
            date            Date     DEFAULT toDate(now()),
            time            String   DEFAULT '',
            channel         String   DEFAULT 'Другой',
            operator        String   DEFAULT '',
            source_type     String   DEFAULT '',
            author          String   DEFAULT '',
            appeal_type     String   DEFAULT '',
            category        String   DEFAULT '',
            subcategory     String   DEFAULT '',
            problem_summary String   DEFAULT '',
            result          String   DEFAULT '',
            comment         String   DEFAULT ''
        ) ENGINE = ReplacingMergeTree(updated_at)
        ORDER BY id
    """)
    for col in ["rating UInt8 DEFAULT 0", "complexity UInt8 DEFAULT 0"]:
        try:
            ch_execute(f"ALTER TABLE manual_log_entries ADD COLUMN IF NOT EXISTS {col}")
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_q = lambda s: s.replace("'", "\\'")


def get_manual_rows(df: str, dt: str, operator: str = "", channel: str = "",
                    stype: str = "", category: str = "", subcategory: str = "",
                    result: str = "", author: str = "", login: str = "",
                    search: str = "", dept: str = "") -> list:
    """Загружает ручные строки из manual_log_entries за период."""
    try:
        where = f"toDate(date) BETWEEN '{df}' AND '{dt}'"
        if operator:    where += f" AND operator    = '{_q(operator)}'"
        if channel:     where += f" AND channel     = '{_q(channel)}'"
        if stype:       where += f" AND source_type = '{_q(stype)}'"
        if category:    where += f" AND category    = '{_q(category)}'"
        if subcategory: where += f" AND subcategory = '{_q(subcategory)}'"
        if result:      where += f" AND result      = '{_q(result)}'"
        # Фильтры которых нет у ручных строк — исключаем все ручные строки
        if dept:        return []   # у ручных строк нет отдела
        if login and len(login) >= 3: return []  # у ручных строк нет логина
        if author:      where += f" AND author ILIKE '%{_q(author)}%'"
        if search and len(search) >= 3:
            sq = _q(search)
            where += (
                f" AND (author ILIKE '%{sq}%'"
                f" OR operator ILIKE '%{sq}%'"
                f" OR problem_summary ILIKE '%{sq}%'"
                f" OR category ILIKE '%{sq}%'"
                f" OR subcategory ILIKE '%{sq}%'"
                f" OR result ILIKE '%{sq}%'"
                f" OR comment ILIKE '%{sq}%')"
            )

        rows = ch_query(f"""
            SELECT
                concat('m_', toString(id)) AS row_key,
                0                          AS chat_id,
                toString(date)             AS date,
                time, operator, source_type, author,
                ''                         AS login,
                appeal_type, category, subcategory,
                problem_summary, result, comment, channel,
                ifNull(rating,     0)      AS rating,
                ifNull(complexity, 0)      AS complexity
            FROM manual_log_entries FINAL
            WHERE {where}
            ORDER BY date DESC, time DESC
            FORMAT JSONEachRow
        """)
        for r in rows:
            r["is_manual"] = True
        return rows
    except Exception:
        return []

# ---------------------------------------------------------------------------
# Права доступа по группам Authentik
# ---------------------------------------------------------------------------

ALLOWED_GROUPS  = {"communication-leads", "communication-managers", "communication-support"}
MANAGED_GROUPS  = ["communication-leads", "communication-managers", "communication-support"]

# Все колонки журнала: (field_name, display_name)
COLUMN_DEFS = [
    ("date",            "Дата"),
    ("time",            "Время"),
    ("operator",        "Оператор"),
    ("source_type",     "Тип автора"),
    ("author",          "Автор"),
    ("login",           "Логин"),
    ("appeal_type",     "Тип"),
    ("category",        "Категория"),
    ("subcategory",     "Подкатегория"),
    ("problem_summary", "Причина обращения"),
    ("result",          "Результат"),
    ("rating",          "Оценка"),
    ("complexity",      "Сложность"),
    ("comment",         "Комм. поддержки"),
    ("comment_manager", "Комм. менеджера"),
    ("responsible_dept","Отв. отдел"),
    ("channel",         "Канал"),
    ("chat_id",         "№ обращения"),
]

# Права по умолчанию (когда БД пуста)
_DEFAULT_COL_PERMS = {
    "communication-leads": {
        "date": "view", "time": "view", "operator": "view",
        "source_type": "view", "author": "view", "login": "view",
        "appeal_type": "view", "category": "view", "subcategory": "view",
        "problem_summary": "view", "result": "view", "comment": "view",
        "comment_manager": "view", "responsible_dept": "view",
        "channel": "view", "chat_id": "view",
        "rating": "edit", "complexity": "edit",
    },
    "communication-managers": {
        "date": "view", "time": "view", "operator": "view",
        "source_type": "edit", "author": "view", "login": "view",
        "appeal_type": "view", "category": "edit", "subcategory": "edit",
        "problem_summary": "view", "result": "edit", "comment": "view",
        "comment_manager": "edit", "responsible_dept": "edit",
        "channel": "view", "chat_id": "view",
        "rating": "edit", "complexity": "edit",
    },
    "communication-support": {
        "date": "view", "time": "view", "operator": "view",
        "source_type": "edit", "author": "view", "login": "view",
        "appeal_type": "view", "category": "view", "subcategory": "view",
        "problem_summary": "view", "result": "view", "comment": "edit",
        "comment_manager": "view", "responsible_dept": "edit",
        "channel": "view", "chat_id": "view",
        "rating": "view", "complexity": "view",
    },
}


def ensure_col_perms_table():
    ch_execute("""
        CREATE TABLE IF NOT EXISTS group_column_perms (
            group_name   String,
            column_name  String,
            access       String DEFAULT 'view',
            updated_at   DateTime DEFAULT now()
        ) ENGINE = ReplacingMergeTree(updated_at)
        ORDER BY (group_name, column_name)
    """)


def load_col_perms(group_name: str) -> dict:
    """Возвращает {column_name: access} для группы.
    Мёрджит дефолты с БД: БД перекрывает дефолты, новые колонки берутся из дефолтов."""
    result = dict(_DEFAULT_COL_PERMS.get(group_name, {}))
    try:
        rows = ch_query(f"""
            SELECT column_name, access
            FROM group_column_perms FINAL
            WHERE group_name = '{group_name}'
            FORMAT JSONEachRow
        """)
        if rows:
            result.update({r["column_name"]: r["access"] for r in rows})
    except Exception:
        pass
    return result


def get_permissions(user: dict):
    groups = set(user.get("groups", []))
    if not groups & ALLOWED_GROUPS:
        return None  # нет доступа

    # Приоритет: leads > managers > support
    is_lead    = "communication-leads"    in groups
    is_manager = "communication-managers" in groups

    if is_lead:
        group_name = "communication-leads"
        role       = "lead"
    elif is_manager:
        group_name = "communication-managers"
        role       = "manager"
    else:
        group_name = "communication-support"
        role       = "support"

    col_perms = load_col_perms(group_name)
    editable = [c for c, a in col_perms.items() if a == "edit"]
    hidden   = [c for c, a in col_perms.items() if a == "hidden"]

    return {
        "editable":   editable,
        "hidden":     hidden,
        "role":       role,
        "is_manager": is_manager,   # оставляем для обратной совместимости
        "is_lead":    is_lead,
    }


# ---------------------------------------------------------------------------
# API
# ---------------------------------------------------------------------------

PER_PAGE = 100

@router.get("/api/log")
def api_log(
    date_from:   str = Query(default=None),
    date_to:     str = Query(default=None),
    operator:    str = Query(default=""),
    channel:     str = Query(default=""),
    stype:       str = Query(default=""),
    category:    str = Query(default=""),
    subcategory: str = Query(default=""),
    result:      str = Query(default=""),
    dept:        str = Query(default=""),
    author:      str = Query(default=""),
    login:       str = Query(default=""),
    search:      str = Query(default=""),
    page:        int = Query(default=1, ge=1),
):
    df = date_from or str(date.today() - timedelta(days=7))
    dt = date_to   or str(date.today())

    def q(s): return s.replace("'", "\\'")

    where = f"toDate(d.event_timestamp) BETWEEN '{df}' AND '{dt}'"
    if operator:
        where += f" AND d.operator_name = '{q(operator)}'"
    if channel == "Чат":
        where += " AND d.source = 'jivo'"
    elif channel == "ЛС":
        where += " AND d.source = 'site_pm'"
    elif channel == "Жалоба":
        where += " AND d.source = 'claim'"
    elif channel:
        where += " AND 1=0"  # любой другой канал — только ручные строки
    # Фильтры по вычисляемым полям (с учётом правок из support_log_edits)
    if stype:
        where += f" AND if(e.source_type!=''  , e.source_type , ifNull(a.source_type,''))  = '{q(stype)}'"
    if category:
        where += f" AND if(e.category!=''     , e.category    , ifNull(a.category,''))     = '{q(category)}'"
    if subcategory:
        where += f" AND if(e.subcategory!=''  , e.subcategory , ifNull(a.subcategory,''))  = '{q(subcategory)}'"
    if result:
        where += f" AND if(e.result!='' AND e.result IS NOT NULL, e.result, ifNull(a.resolution_status,'')) = '{q(result)}'"
    if dept:
        where += f" AND ifNull(t.responsible_dept, '') = '{q(dept)}'"
    if author:
        where += f" AND ifNull(d.visitor_name, '') ILIKE '%{q(author)}%'"
    if login and len(login) >= 3:
        where += f" AND ifNull(d.visitor_name, '') ILIKE '%{q(login)}%'"
    if search and len(search) >= 3:
        sq = q(search)
        where += (
            f" AND ("
            f"ifNull(d.visitor_name,'') ILIKE '%{sq}%'"
            f" OR ifNull(d.operator_name,'') ILIKE '%{sq}%'"
            f" OR ifNull(a.contact_reason,'') ILIKE '%{sq}%'"
            f" OR if(e.category!='',e.category,ifNull(a.category,'')) ILIKE '%{sq}%'"
            f" OR if(e.subcategory!='',e.subcategory,ifNull(a.subcategory,'')) ILIKE '%{sq}%'"
            f" OR ifNull(a.user_problem_summary,'') ILIKE '%{sq}%'"
            f" OR if(e.result!='' AND e.result IS NOT NULL,e.result,ifNull(a.resolution_status,'')) ILIKE '%{sq}%'"
            f" OR ifNull(e.comment,'') ILIKE '%{sq}%'"
            f" OR ifNull(e.comment_manager,'') ILIKE '%{sq}%'"
            f" OR ifNull(t.responsible_dept,'') ILIKE '%{sq}%'"
            f" OR ifNull(d.page_url,'') ILIKE '%{sq}%'"
            f")"
        )

    offset = (page - 1) * PER_PAGE

    # Общее количество для пагинации
    total_rows = ch_query(f"""
        SELECT count() AS total
        FROM dialogs d
        LEFT JOIN dialog_analysis a ON d.chat_id = a.chat_id AND d.source = a.source
        LEFT JOIN (SELECT * FROM support_log_edits FINAL) e ON d.chat_id = e.chat_id
        LEFT JOIN (SELECT * FROM day_tracker_edits FINAL) t ON d.chat_id = t.chat_id
        WHERE {where}
        FORMAT JSONEachRow
    """)
    total = int((total_rows[0] if total_rows else {}).get("total", 0))

    rows = ch_query(f"""
        SELECT
            d.chat_id                                                              AS chat_id,
            toString(d.chat_id)                                                    AS row_key,
            toString(toDate(d.event_timestamp))                                    AS date,
            substring(toString(d.event_timestamp), 12, 5)                         AS time,
            ifNull(d.operator_name, '')                                            AS operator,
            ifNull(d.visitor_name, '')                                             AS author,
            toString(ifNull(d.visitor_id, 0))                                      AS login,
            ifNull(a.contact_reason, '')                                           AS appeal_type,
            if(e.source_type  != '', e.source_type,  ifNull(a.source_type, ''))   AS source_type,
            if(e.category     != '', e.category,     ifNull(a.category, ''))      AS category,
            if(e.subcategory  != '', e.subcategory,  ifNull(a.subcategory, ''))   AS subcategory,
            ifNull(a.user_problem_summary, '')                                     AS problem_summary,
            if(e.result IS NOT NULL AND e.result != '',
               e.result,
               ifNull(a.resolution_status, ''))                                    AS result,
            ifNull(e.comment, '')                                                  AS comment,
            ifNull(e.comment_author, '')                                           AS comment_author,
            ifNull(e.comment_manager, '')                                          AS comment_manager,
            ifNull(e.comment_manager_author, '')                                   AS comment_manager_author,
            ifNull(t.responsible_dept, '')                                         AS responsible_dept,
            multiIf(d.source='jivo','Чат',d.source='site_pm','ЛС',d.source='claim','Жалоба',d.source) AS channel,
            ifNull(e.rating,     0)                                                AS rating,
            ifNull(e.complexity, 0)                                                AS complexity
        FROM dialogs d
        LEFT JOIN dialog_analysis a ON d.chat_id = a.chat_id AND d.source = a.source
        LEFT JOIN (SELECT * FROM support_log_edits FINAL) e ON d.chat_id = e.chat_id
        LEFT JOIN (SELECT * FROM day_tracker_edits FINAL) t ON d.chat_id = t.chat_id
        WHERE {where}
        ORDER BY d.event_timestamp DESC
        LIMIT {PER_PAGE} OFFSET {offset}
        FORMAT JSONEachRow
    """)
    for r in rows:
        r["is_manual"] = False

    # Ручные строки — подмешиваем на первой странице и сортируем вместе с основными
    manual_rows = []
    if page == 1:
        manual_rows = get_manual_rows(df, dt, operator, channel, stype, category, subcategory, result,
                                      author=author, login=login, search=search, dept=dept)
    all_rows = sorted(
        manual_rows + rows,
        key=lambda r: (r.get("date", ""), r.get("time", "")),
        reverse=True,
    )

    operators = ch_query("""
        SELECT DISTINCT operator_name
        FROM dialogs
        WHERE operator_name != '' AND operator_name IS NOT NULL
        ORDER BY operator_name
        FORMAT JSONEachRow
    """)

    # Каналы — из диалогов + из ручных строк
    ch_sources = ch_query("""
        SELECT multiIf(source='jivo','Чат',source='site_pm','ЛС',source='claim','Жалоба',source) AS channel
        FROM dialogs
        GROUP BY source
        FORMAT JSONEachRow
    """)
    manual_channels = ch_query("""
        SELECT DISTINCT channel FROM manual_log_entries
        WHERE channel != '' AND channel IS NOT NULL
        ORDER BY channel
        FORMAT JSONEachRow
    """)
    channels = sorted({r["channel"] for r in ch_sources + manual_channels if r.get("channel")})

    manual_total = len(get_manual_rows(df, dt, operator, channel, stype, category, subcategory, result,
                                       author=author, login=login, search=search, dept=dept)) if page != 1 else len(manual_rows)

    return JSONResponse({
        "rows":      all_rows,
        "operators": [r["operator_name"] for r in operators],
        "channels":  channels,
        "total":     total + manual_total,
        "page":      page,
        "per_page":  PER_PAGE,
        "pages":     max(1, -(-total // PER_PAGE)),  # ceil division
    })


@router.get("/api/log/dialog/{chat_id}")
def api_dialog(chat_id: int):
    rows = ch_query(f"""
        SELECT chat_messages_json,
               ifNull(visitor_name, '')  AS visitor_name,
               ifNull(operator_name, '') AS operator_name
        FROM dialogs
        WHERE chat_id = {chat_id}
        LIMIT 1
        FORMAT JSONEachRow
    """)
    if not rows:
        return JSONResponse({"messages": [], "visitor": "", "operator": ""})
    r = rows[0]
    try:
        messages = json.loads(r.get("chat_messages_json") or "[]")
    except Exception:
        messages = []
    return JSONResponse({
        "messages": messages,
        "visitor":  r.get("visitor_name", ""),
        "operator": r.get("operator_name", ""),
    })


# ---------------------------------------------------------------------------
# Export to Excel
# ---------------------------------------------------------------------------

@router.get("/api/log/export")
def api_log_export(
    date_from:   str = Query(default=None),
    date_to:     str = Query(default=None),
    operator:    str = Query(default=""),
    channel:     str = Query(default=""),
    stype:       str = Query(default=""),
    category:    str = Query(default=""),
    subcategory: str = Query(default=""),
    result:      str = Query(default=""),
    dept:        str = Query(default=""),
    author:      str = Query(default=""),
    login:       str = Query(default=""),
    search:      str = Query(default=""),
    user: dict = Depends(require_user),
):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return JSONResponse({"error": "openpyxl не установлен: pip install openpyxl"}, status_code=500)

    df = date_from or str(date.today() - timedelta(days=7))
    dt = date_to   or str(date.today())

    def q(s): return s.replace("'", "\\'")

    where = f"toDate(d.event_timestamp) BETWEEN '{df}' AND '{dt}'"
    if operator:    where += f" AND d.operator_name = '{q(operator)}'"
    if channel == "Чат":    where += " AND d.source = 'jivo'"
    elif channel == "ЛС":   where += " AND d.source = 'site_pm'"
    elif channel == "Жалоба": where += " AND d.source = 'claim'"
    elif channel:           where += " AND 1=0"
    if stype:       where += f" AND if(e.source_type!='',e.source_type,ifNull(a.source_type,''))='{q(stype)}'"
    if category:    where += f" AND if(e.category!='',e.category,ifNull(a.category,''))='{q(category)}'"
    if subcategory: where += f" AND if(e.subcategory!='',e.subcategory,ifNull(a.subcategory,''))='{q(subcategory)}'"
    if result:      where += f" AND if(e.result!='' AND e.result IS NOT NULL,e.result,ifNull(a.resolution_status,''))='{q(result)}'"
    if dept:        where += f" AND ifNull(t.responsible_dept,'')='{q(dept)}'"
    if author:      where += f" AND ifNull(d.visitor_name,'') ILIKE '%{q(author)}%'"
    if login and len(login) >= 3: where += f" AND ifNull(d.visitor_name,'') ILIKE '%{q(login)}%'"
    if search and len(search) >= 3:
        sq = q(search)
        where += (
            f" AND (ifNull(d.visitor_name,'') ILIKE '%{sq}%'"
            f" OR ifNull(d.operator_name,'') ILIKE '%{sq}%'"
            f" OR ifNull(a.user_problem_summary,'') ILIKE '%{sq}%'"
            f" OR if(e.category!='',e.category,ifNull(a.category,'')) ILIKE '%{sq}%'"
            f")"
        )

    rows = ch_query(f"""
        SELECT
            d.chat_id                                                              AS chat_id,
            toString(toDate(d.event_timestamp))                                    AS date,
            substring(toString(d.event_timestamp), 12, 5)                         AS time,
            ifNull(d.operator_name, '')                                            AS operator,
            ifNull(d.visitor_name, '')                                             AS author,
            toString(ifNull(d.visitor_id, 0))                                      AS login,
            ifNull(a.contact_reason, '')                                           AS appeal_type,
            if(e.source_type  != '', e.source_type,  ifNull(a.source_type, ''))   AS source_type,
            if(e.category     != '', e.category,     ifNull(a.category, ''))      AS category,
            if(e.subcategory  != '', e.subcategory,  ifNull(a.subcategory, ''))   AS subcategory,
            ifNull(a.user_problem_summary, '')                                     AS problem_summary,
            if(e.result IS NOT NULL AND e.result != '',
               e.result, ifNull(a.resolution_status, ''))                          AS result,
            ifNull(e.comment, '')                                                  AS comment,
            ifNull(e.comment_manager, '')                                          AS comment_manager,
            ifNull(t.responsible_dept, '')                                         AS responsible_dept,
            multiIf(d.source='jivo','Чат',d.source='site_pm','ЛС',d.source='claim','Жалоба',d.source) AS channel,
            ifNull(e.rating,     0)                                                AS rating,
            ifNull(e.complexity, 0)                                                AS complexity,
            ifNull(d.plain_messages, '')                                           AS dialog_text
        FROM dialogs d
        LEFT JOIN dialog_analysis a ON d.chat_id = a.chat_id AND d.source = a.source
        LEFT JOIN (SELECT * FROM support_log_edits FINAL) e ON d.chat_id = e.chat_id
        LEFT JOIN (SELECT * FROM day_tracker_edits FINAL) t ON d.chat_id = t.chat_id
        WHERE {where}
        ORDER BY d.event_timestamp DESC
        LIMIT 10000
        FORMAT JSONEachRow
    """)

    # Ручные строки
    manual = get_manual_rows(df, dt, operator, channel, stype, category, subcategory, result,
                             author=author, login=login, search=search, dept=dept)

    all_rows = sorted(rows + manual, key=lambda r: (r.get("date",""), r.get("time","")), reverse=True)

    # ── Сборка Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Обращения"

    HEADERS = [
        ("Дата",              12), ("Время",        8),  ("Оператор",         18),
        ("Тип автора",        14), ("Автор",        22), ("Логин",            16),
        ("Тип обращения",     16), ("Категория",    22), ("Подкатегория",     26),
        ("Причина обращения", 40), ("Результат",    14), ("Оценка",           8),
        ("Сложность",         10), ("Комм. поддержки", 30), ("Комм. менеджера", 30),
        ("Отв. отдел",        16), ("Канал",        12), ("№ обращения",      14),
        ("Диалог",            60),
    ]

    # Стиль заголовка
    hdr_fill = PatternFill("solid", fgColor="1A73E8")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, (hdr, width) in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font       = hdr_font
        cell.fill       = hdr_fill
        cell.alignment  = hdr_align
        cell.border     = border
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # Стили строк
    fill_even = PatternFill("solid", fgColor="F8FAFF")
    cell_align = Alignment(vertical="top", wrap_text=True)
    cell_align_center = Alignment(horizontal="center", vertical="top")

    for row_idx, r in enumerate(all_rows, 2):
        fill = fill_even if row_idx % 2 == 0 else None
        is_manual = r.get("is_manual", False)

        def fmt_rating(val):
            v = int(val or 0)
            return str(v) if v else ""

        values = [
            r.get("date", "").replace("-", ".") if "-" in r.get("date","") else r.get("date",""),
            r.get("time", ""),
            r.get("operator", ""),
            r.get("source_type", ""),
            r.get("author", ""),
            r.get("login", ""),
            r.get("appeal_type", ""),
            r.get("category", ""),
            r.get("subcategory", ""),
            r.get("problem_summary", ""),
            r.get("result", ""),
            fmt_rating(r.get("rating", 0)),
            fmt_rating(r.get("complexity", 0)),
            r.get("comment", ""),
            r.get("comment_manager", ""),
            r.get("responsible_dept", ""),
            r.get("channel", ""),
            "" if is_manual else str(r.get("chat_id", "")),
            r.get("dialog_text", "") if not is_manual else "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = cell_align_center if col_idx in (1, 2, 12, 13, 18) else cell_align
            if fill:
                cell.fill = fill
            if is_manual:
                cell.font = Font(color="888888", size=9)
            else:
                cell.font = Font(size=9)

        # Высота строки — побольше если есть диалог
        dialog = values[-1]
        lines = min(dialog.count("\n") + 1, 30) if dialog else 1
        ws.row_dimensions[row_idx].height = max(16, min(lines * 13, 300))

    # Итого строк
    ws.cell(row=len(all_rows) + 3, column=1,
            value=f"Итого: {len(all_rows)} обращений").font = Font(italic=True, color="888888", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"обращения_{df}_{dt}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{urllib.parse.quote(fname)}"},
    )


@router.post("/api/log/manual")
def api_create_manual():
    """Создаёт новую пустую строку вручную."""
    ensure_manual_table()
    from datetime import datetime as _dt
    now    = _dt.now()
    new_id = int(now.timestamp() * 1000)
    today  = str(now.date())
    now_time = now.strftime("%H:%M")
    ch_execute(
        "INSERT INTO manual_log_entries (id, date, time, channel) FORMAT JSONEachRow",
        data=json.dumps({"id": new_id, "date": today, "time": now_time, "channel": ""}).encode("utf-8"),
    )
    return JSONResponse({"ok": True, "id": new_id, "date": today, "time": now_time})


class ManualEditPayload(BaseModel):
    date:            Optional[str] = ""
    time:            Optional[str] = ""
    channel:         Optional[str] = ""
    operator:        Optional[str] = ""
    source_type:     Optional[str] = ""
    author:          Optional[str] = ""
    appeal_type:     Optional[str] = ""
    category:        Optional[str] = ""
    subcategory:     Optional[str] = ""
    problem_summary: Optional[str] = ""
    result:          Optional[str] = ""
    comment:         Optional[str] = ""
    rating:          Optional[int] = None
    complexity:      Optional[int] = None


@router.delete("/api/log/manual/{row_id}")
def api_delete_manual(row_id: int):
    ch_execute(f"ALTER TABLE manual_log_entries DELETE WHERE id = {row_id}")
    return JSONResponse({"ok": True})


@router.post("/api/log/manual/{row_id}")
def api_edit_manual(row_id: int, payload: ManualEditPayload):
    today = str(date.today())
    row = json.dumps({
        "id":            row_id,
        "date":          payload.date or today,
        "time":          payload.time or "",
        "channel":       payload.channel or "",
        "operator":      payload.operator or "",
        "source_type":   payload.source_type or "",
        "author":        payload.author or "",
        "appeal_type":   payload.appeal_type or "",
        "category":      payload.category or "",
        "subcategory":   payload.subcategory or "",
        "problem_summary": payload.problem_summary or "",
        "result":        payload.result or "",
        "comment":       payload.comment or "",
        "rating":        int(payload.rating)     if payload.rating     is not None else 0,
        "complexity":    int(payload.complexity) if payload.complexity is not None else 0,
    }, ensure_ascii=False)
    ch_execute(
        "INSERT INTO manual_log_entries "
        "(id, date, time, channel, operator, source_type, author, appeal_type, "
        "category, subcategory, problem_summary, result, comment, rating, complexity) "
        "FORMAT JSONEachRow",
        data=row.encode("utf-8"),
    )
    return JSONResponse({"ok": True})


class EditPayload(BaseModel):
    organizer:                Optional[str] = ""
    responsible:              Optional[str] = ""
    result:                   Optional[str] = ""
    comment:                  Optional[str] = ""
    comment_author:           Optional[str] = ""
    comment_manager:          Optional[str] = ""
    comment_manager_author:   Optional[str] = ""
    source_type:              Optional[str] = ""
    category:                 Optional[str] = ""
    subcategory:              Optional[str] = ""
    responsible_dept:         Optional[str] = None   # None = поле не пришло → не трогаем day_tracker_edits
    rating:                   Optional[int] = None   # None = не менялось
    complexity:               Optional[int] = None   # None = не менялось


@router.post("/api/log/{chat_id}")
def api_edit(chat_id: int, payload: EditPayload):
    row = json.dumps({
        "chat_id":                  chat_id,
        "organizer":                payload.organizer               or "",
        "responsible":              payload.responsible             or "",
        "result":                   payload.result                  or "",
        "comment":                  payload.comment                 or "",
        "comment_author":           payload.comment_author          or "",
        "comment_manager":          payload.comment_manager         or "",
        "comment_manager_author":   payload.comment_manager_author  or "",
        "source_type":              payload.source_type             or "",
        "category":                 payload.category                or "",
        "subcategory":              payload.subcategory             or "",
        "rating":                   int(payload.rating)     if payload.rating     is not None else 0,
        "complexity":               int(payload.complexity) if payload.complexity is not None else 0,
    }, ensure_ascii=False)
    ch_execute(
        "INSERT INTO support_log_edits "
        "(chat_id, organizer, responsible, result, comment, comment_author, "
        "comment_manager, comment_manager_author, source_type, category, subcategory, "
        "rating, complexity) "
        "FORMAT JSONEachRow",
        data=row.encode("utf-8"),
    )
    if payload.responsible_dept is not None:
        dept_row = json.dumps({
            "chat_id":          chat_id,
            "responsible_dept": payload.responsible_dept or "",
        }, ensure_ascii=False)
        ch_execute(
            "INSERT INTO day_tracker_edits (chat_id, responsible_dept) FORMAT JSONEachRow",
            data=dept_row.encode("utf-8"),
        )
    return JSONResponse({"ok": True})


# ---------------------------------------------------------------------------
# Admin: права доступа по колонкам
# ---------------------------------------------------------------------------

@router.get("/api/admin/perms")
def api_admin_perms_get(user: dict = Depends(require_user)):
    if "communication-leads" not in set(user.get("groups", [])):
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    result = {g: load_col_perms(g) for g in MANAGED_GROUPS}
    return JSONResponse(result)


@router.post("/api/admin/perms")
async def api_admin_perms_save(request: Request, user: dict = Depends(require_user)):
    if "communication-leads" not in set(user.get("groups", [])):
        return JSONResponse({"error": "Forbidden"}, status_code=403)
    payload = await request.json()
    valid_cols  = {c for c, _ in COLUMN_DEFS}
    valid_accs  = {"view", "edit", "hidden"}
    rows = []
    for group_name, cols in payload.items():
        if group_name not in MANAGED_GROUPS:
            continue
        for col_name, access in cols.items():
            if col_name not in valid_cols or access not in valid_accs:
                continue
            rows.append(json.dumps({
                "group_name":  group_name,
                "column_name": col_name,
                "access":      access,
            }, ensure_ascii=False))
    if rows:
        ch_execute(
            "INSERT INTO group_column_perms (group_name, column_name, access) FORMAT JSONEachRow",
            data="\n".join(rows).encode("utf-8"),
        )
    return JSONResponse({"ok": True})


@router.get("/admin/perms", response_class=HTMLResponse)
def admin_perms_page(user: dict = Depends(require_user)):
    if "communication-leads" not in set(user.get("groups", [])):
        from fastapi.responses import HTMLResponse as _HR
        return _HR("<h2 style='font-family:sans-serif;margin:60px auto;text-align:center'>403 — Доступ запрещён</h2>", status_code=403)
    col_labels   = {f: lbl for f, lbl in COLUMN_DEFS}
    group_labels = {
        "communication-leads":    "Руководители",
        "communication-managers": "Менеджеры",
        "communication-support":  "Поддержка",
    }
    return HTMLResponse(_build_admin_html(col_labels, group_labels))


def _build_admin_html(col_labels: dict, group_labels: dict) -> str:
    access_opts = [("view", "Просмотр"), ("edit", "Редактировать"), ("hidden", "Скрыть")]

    # Строим строки таблицы
    rows_html = ""
    for field, label in COLUMN_DEFS:
        cells = f'<td class="col-label">{label}</td>'
        for gname in MANAGED_GROUPS:
            sel_id = f"perm_{gname}__{field}"
            opts = "".join(
                f'<option value="{v}">{lbl}</option>'
                for v, lbl in access_opts
            )
            cells += f'<td><select id="{sel_id}" data-group="{gname}" data-col="{field}">{opts}</select></td>'
        rows_html += f"<tr>{cells}</tr>"

    group_ths = "".join(f"<th>{group_labels[g]}</th>" for g in MANAGED_GROUPS)

    return f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<title>Настройка прав колонок</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
       font-size: 13px; background: #f0f2f5; color: #222; padding: 24px; }}
h1 {{ font-size: 18px; margin-bottom: 20px; }}
.card {{ background: #fff; border-radius: 8px; box-shadow: 0 1px 4px rgba(0,0,0,.1);
         padding: 24px; max-width: 680px; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ padding: 8px 12px; border-bottom: 1px solid #eee; text-align: left; }}
th {{ background: #f8f9fa; font-weight: 600; color: #555; font-size: 12px; text-transform: uppercase; }}
td.col-label {{ font-weight: 500; color: #333; width: 200px; }}
select {{
  padding: 4px 8px; border: 1px solid #ddd; border-radius: 5px;
  font-size: 12px; background: #fff; cursor: pointer; width: 150px;
  transition: border-color .15s;
}}
select:focus {{ outline: none; border-color: #1a73e8; }}
select[value="edit"]   {{ border-color: #27ae60; color: #27ae60; }}
select[value="hidden"] {{ border-color: #e74c3c; color: #e74c3c; }}
.btn-save {{
  margin-top: 20px; padding: 9px 28px; background: #1a73e8; color: #fff;
  border: none; border-radius: 6px; font-size: 14px; cursor: pointer;
  transition: background .15s;
}}
.btn-save:hover {{ background: #1558b0; }}
.btn-save:disabled {{ background: #aaa; cursor: default; }}
.status {{ margin-top: 12px; font-size: 13px; color: #27ae60; min-height: 18px; }}
.back {{ display: inline-block; margin-bottom: 16px; font-size: 13px;
         color: #1a73e8; text-decoration: none; }}
.back:hover {{ text-decoration: underline; }}
</style>
</head>
<body>
<a href="/log" class="back">← Журнал обращений</a>
<h1>Настройка прав доступа к колонкам</h1>
<div class="card">
  <table>
    <thead><tr><th>Колонка</th>{group_ths}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  <button class="btn-save" id="btn_save" onclick="savePerms()">Сохранить</button>
  <div class="status" id="status"></div>
</div>
<script>
async function loadPerms() {{
  const resp = await fetch('/api/admin/perms');
  const data = await resp.json();
  document.querySelectorAll('select[data-group]').forEach(sel => {{
    const g = sel.dataset.group, c = sel.dataset.col;
    if (data[g] && data[g][c]) {{
      sel.value = data[g][c];
      applySelectColor(sel);
    }}
  }});
}}

function applySelectColor(sel) {{
  sel.style.borderColor = sel.value === 'edit' ? '#27ae60'
                        : sel.value === 'hidden' ? '#e74c3c' : '#ddd';
  sel.style.color = sel.value === 'edit' ? '#27ae60'
                  : sel.value === 'hidden' ? '#e74c3c' : '#222';
}}

document.querySelectorAll('select[data-group]').forEach(sel => {{
  sel.addEventListener('change', () => applySelectColor(sel));
}});

async function savePerms() {{
  const btn = document.getElementById('btn_save');
  const st  = document.getElementById('status');
  btn.disabled = true;
  st.textContent = '';

  const payload = {{}};
  document.querySelectorAll('select[data-group]').forEach(sel => {{
    const g = sel.dataset.group, c = sel.dataset.col;
    if (!payload[g]) payload[g] = {{}};
    payload[g][c] = sel.value;
  }});

  try {{
    const resp = await fetch('/api/admin/perms', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: JSON.stringify(payload),
    }});
    const data = await resp.json();
    st.style.color = data.ok ? '#27ae60' : '#e74c3c';
    st.textContent = data.ok ? '✓ Сохранено. Изменения вступят в силу при следующем входе пользователей.' : ('Ошибка: ' + JSON.stringify(data));
  }} catch(e) {{
    st.style.color = '#e74c3c';
    st.textContent = 'Ошибка: ' + e.message;
  }}
  btn.disabled = false;
}}

loadPerms();
</script>
</body>
</html>"""


# ---------------------------------------------------------------------------
# HTML
# ---------------------------------------------------------------------------

@router.get("/log", response_class=HTMLResponse)
def log_page(user: dict = Depends(require_user)):
    perms = get_permissions(user)
    if perms is None:
        from fastapi.responses import HTMLResponse as _HR
        name = user.get("name") or user.get("username", "")
        return _HR(
            f"<div style='font-family:sans-serif;margin:80px auto;text-align:center'>"
            f"<h2>403 — Доступ запрещён</h2>"
            f"<p style='font-size:14px;color:#888;margin:12px 0'>Пользователь <b>{name}</b> не состоит "
            f"в группах communication-leads, communication-managers или communication-support</p>"
            f"<p style='font-size:13px;color:#aaa;margin:6px 0'>Если вас только что добавили в группу — "
            f"выйдите и войдите снова, чтобы сессия обновилась.</p>"
            f"<a href='/auth/logout' style='display:inline-block;margin-top:18px;padding:8px 22px;"
            f"background:#e74c3c;color:#fff;border-radius:6px;text-decoration:none;font-size:14px'>"
            f"Выйти и войти снова</a>"
            f"</div>",
            status_code=403,
        )
    admin_link = (
        '<a href="/admin/perms" class="topbar-admin">⚙ Права</a>'
        if perms.get("is_lead") else ""
    )
    return (
        _HTML
        .replace("%%USER%%",       user.get("name") or user.get("username", ""))
        .replace("%%PERMS%%",      json.dumps(perms, ensure_ascii=False))
        .replace("%%ADMIN_LINK%%", admin_link)
    )


@router.get("/day-tracker", response_class=HTMLResponse)
def day_tracker_redirect():
    from fastapi.responses import RedirectResponse
    return RedirectResponse("/log", status_code=302)


_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Журнал обращений</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
html, body {
  height: 100%; overflow: hidden;
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  font-size: 13px; background: #f0f2f5; color: #222;
  display: flex; flex-direction: column;
}

/* ── Topbar ── */
.topbar {
  flex-shrink: 0; background: #1a73e8;
  padding: 0 16px; height: 48px;
  display: flex; align-items: center; gap: 10px;
  box-shadow: 0 2px 6px rgba(0,0,0,.18); z-index: 10;
}
.topbar h1 {
  font-size: 15px; font-weight: 600; color: #fff;
  white-space: nowrap; letter-spacing: .2px;
}
.topbar-right { margin-left: auto; display: flex; align-items: center; gap: 4px; }
.topbar-search { position: relative; display: flex; align-items: center; }
.topbar-search input {
  border: none; border-radius: 20px;
  padding: 5px 12px 5px 30px;
  font-size: 12px; background: rgba(255,255,255,.18); color: #fff;
  width: 200px; outline: none; transition: background .2s, width .2s;
}
.topbar-search input::placeholder { color: rgba(255,255,255,.6); }
.topbar-search input:focus { background: rgba(255,255,255,.28); width: 240px; }
.topbar-search-icon {
  position: absolute; left: 9px; font-size: 12px;
  color: rgba(255,255,255,.7); pointer-events: none;
}
.user-chip {
  display: flex; align-items: center; gap: 6px;
  background: rgba(255,255,255,.18); border-radius: 20px;
  padding: 3px 10px 3px 4px; font-size: 12px; color: #fff; white-space: nowrap;
}
.user-avatar {
  width: 26px; height: 26px; border-radius: 50%;
  background: rgba(255,255,255,.35);
  display: flex; align-items: center; justify-content: center;
  font-size: 11px; font-weight: 700; color: #fff;
}
.topbar-link {
  font-size: 12px; color: rgba(255,255,255,.85); text-decoration: none;
  white-space: nowrap; padding: 5px 10px; border-radius: 5px; transition: background .15s;
}
.topbar-link:hover { background: rgba(255,255,255,.18); color: #fff; }
.topbar-admin {
  font-size: 12px; color: #ffd54f; text-decoration: none; font-weight: 500;
  white-space: nowrap; padding: 5px 10px; border-radius: 5px; transition: background .15s;
}
.topbar-admin:hover { background: rgba(255,255,255,.15); }
/* ── Filterbar ── */
.filterbar {
  flex-shrink: 0; background: #f8f9fa; border-bottom: 1px solid #e0e0e0;
  padding: 7px 16px; display: flex; align-items: center; flex-wrap: wrap; gap: 6px;
}
.fg { display: flex; align-items: center; gap: 4px; }
.fg label {
  font-size: 10.5px; color: #888; white-space: nowrap;
  font-weight: 600; text-transform: uppercase; letter-spacing: .4px;
}
.filterbar input[type=date], .filterbar input[type=text] {
  border: 1px solid #d5d5d5; border-radius: 6px; padding: 4px 8px;
  font-size: 12px; background: #fff; color: #333;
  height: 30px; box-sizing: border-box; transition: border-color .15s, box-shadow .15s;
}
.filterbar select {
  border: 1px solid #d5d5d5; border-radius: 6px; padding: 4px 26px 4px 9px;
  font-size: 12px; background: #fff; color: #333;
  height: 30px; box-sizing: border-box; transition: border-color .15s, box-shadow .15s;
  appearance: none; -webkit-appearance: none;
  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23999'/%3E%3C/svg%3E");
  background-repeat: no-repeat; background-position: right 8px center;
  cursor: pointer;
}
.filterbar input[type=date]:focus, .filterbar input[type=text]:focus, .filterbar select:focus {
  outline: none; border-color: #1a73e8; box-shadow: 0 0 0 2px rgba(26,115,232,.12);
}
.filterbar select:hover { border-color: #b0b0b0; }
.filterbar select option { padding: 6px; }
.filter-sep { width: 1px; height: 22px; background: #d8d8d8; margin: 0 2px; }
.filterbar-actions { display: flex; align-items: center; gap: 8px; margin-left: auto; }
.filterbar-count { font-size: 12px; color: #999; white-space: nowrap; padding-left: 4px; border-left: 1px solid #e0e0e0; }
.count { font-size: 12px; color: #888; white-space: nowrap; }
.btn {
  background: #1a73e8; color: #fff; border: none; border-radius: 6px;
  padding: 0 14px; height: 30px; cursor: pointer; font-size: 12px; font-weight: 500;
  white-space: nowrap; transition: background .15s;
}
.btn:hover { background: #1557b0; }

/* ── Table wrapper ── */
.wrap {
  flex: 1;
  overflow: auto;   /* скролл обоих осей внутри контейнера */
  padding: 0 0 10px 0;  /* без отступа сверху — заголовки вплотную к тулбару */
}

table {
  border-collapse: collapse;
  width: max-content; min-width: 100%;
  background: #fff;
  box-shadow: 0 1px 4px rgba(0,0,0,.08);
}

thead th {
  background: #f1f3f4;
  border-bottom: 2px solid #ddd;
  border-right: 1px solid #e0e0e0;
  padding: 8px 10px;
  text-align: left;
  font-weight: 600;
  font-size: 12px;
  color: #444;
  white-space: nowrap;
  position: sticky;
  top: 0;           /* прилипает к верху .wrap, а не вьюпорта */
  z-index: 10;
}
thead th:last-child { border-right: none; }

tbody tr { border-bottom: 1px solid #f0f0f0; }
tbody tr:last-child { border-bottom: none; }
tbody tr:hover { background: #f8fbff; }

tbody td {
  padding: 6px 10px;
  border-right: 1px solid #f0f0f0;
  vertical-align: top;
}
tbody td:last-child { border-right: none; }

/* Column widths */
.col-date    { white-space: nowrap; }
.col-time    { white-space: nowrap; color: #888; }
.col-channel { white-space: nowrap; font-weight: 500; }
.col-author  { max-width: 90px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }

/* ── Ресайз столбцов ── */
.col-resize-handle {
  position: absolute; right: 0; top: 0;
  width: 5px; height: 100%;
  cursor: col-resize;
  user-select: none;
  z-index: 1;
}
.col-resize-handle:hover,
.col-resize-handle.dragging { background: #1a73e8; }
thead th { cursor: grab; }
thead th.th-drag-over { border-left: 2px solid #1a73e8 !important; }
thead th.th-dragging  { opacity: .45; }
.col-summary { max-width: 320px; word-wrap: break-word; line-height: 1.4; }
.col-login   { color: #888; font-size: 12px; }
.col-week    { text-align: center; color: #888; font-size: 12px; }

/* Channel colors */
.ch-chat { color: #1a73e8; }
.ch-ls   { color: #0f9d58; }
.ch-form { color: #e67e22; }

/* Ответственный отдел */
.dept-badge {
  display: inline-block; padding: 2px 8px; border-radius: 10px;
  font-size: 11px; font-weight: 600; white-space: nowrap;
  background: #e0e7ff; color: #3730a3;
}
.select-dept select { border-color: #818cf8 !important; background: #eef2ff !important; }

/* Editable cells */
.editable {
  display: block;
  min-width: 80px;
  min-height: 20px;
  padding: 3px 5px;
  border-radius: 4px;
  outline: none;
  cursor: text;
  white-space: pre-wrap;
  word-break: break-word;
  overflow-wrap: break-word;
  transition: background .15s, box-shadow .15s;
}
.editable:empty::before {
  content: attr(data-placeholder);
  color: #bbb;
  font-style: italic;
  pointer-events: none;
}
.editable:empty { cursor: pointer; }
.editable:focus {
  background: #fffbea;
  box-shadow: 0 0 0 2px #f59e0b55;
}
.editable.saving { opacity: .6; }
.editable.saved  { animation: flash 1s ease forwards; }
.comment-wrap { display: flex; flex-direction: column; gap: 1px; }
.comment-author { font-weight: 600; font-size: 11px; color: #888; line-height: 1.2; }
.comment-ro { font-size: 12px; line-height: 1.4; white-space: pre-wrap; word-break: break-word; }
.btn-append-comment { background: none; border: 1px dashed #ccc; border-radius: 4px; color: #888; font-size: 11px; cursor: pointer; padding: 1px 6px; margin-top: 2px; align-self: flex-start; }
.btn-append-comment:hover { border-color: #888; color: #555; }
.comment-append-area { display: flex; flex-direction: column; gap: 4px; margin-top: 4px; }
.comment-append-input { min-height: 28px; font-size: 12px; padding: 3px 6px; border: 1px solid #ccc; border-radius: 4px; outline: none; white-space: pre-wrap; word-break: break-word; }
.comment-append-input:empty::before { content: attr(data-placeholder); color: #bbb; }
.comment-append-input:focus { border-color: #4a90d9; background: #fffbea; }
.comment-append-actions { display: flex; gap: 6px; }
.btn-app-save { background: #4a90d9; color: #fff; border: none; border-radius: 4px; font-size: 11px; cursor: pointer; padding: 2px 10px; }
.btn-app-save:hover { background: #2a70b9; }
.btn-app-cancel { background: none; border: 1px solid #ccc; border-radius: 4px; font-size: 11px; cursor: pointer; padding: 2px 8px; color: #666; }
mark.hl { background: #ffe566; color: inherit; border-radius: 2px; padding: 0 1px; font-style: normal; }
.select-cell {
  cursor: pointer; padding: 3px 7px; border-radius: 5px;
  min-width: 60px; min-height: 20px;
  transition: background .15s;
  display: flex; align-items: center; gap: 4px;
}
.select-cell:hover { background: #f0f7ff; }
.select-cell::after {
  content: '▾'; font-size: 10px; color: #bbb; flex-shrink: 0;
  transition: color .15s;
}
.select-cell:hover::after { color: #1a73e8; }
.select-cell.saved { animation: flash 1s ease forwards; }
/* Cell dropdown overlay */
.cell-dd {
  position: fixed; z-index: 9999;
  background: #fff; border: 1px solid #e0e0e0; border-radius: 10px;
  box-shadow: 0 8px 32px rgba(0,0,0,.14); padding: 4px 0;
  min-width: 140px; max-height: 260px; overflow-y: auto;
}
.cell-dd-opt {
  padding: 8px 16px; font-size: 13px; color: #333;
  cursor: pointer; transition: background .1s;
  display: flex; align-items: center; gap: 8px;
}
.cell-dd-opt:hover { background: #f0f7ff; color: #1a73e8; }
.cell-dd-opt.active {
  background: #e8f0fe; color: #1a73e8; font-weight: 600;
}
.cell-dd-opt.active::before {
  content: '✓'; font-size: 11px; color: #1a73e8;
}
@keyframes flash {
  0%   { background: #d1fae5; }
  100% { background: transparent; }
}

/* Result badge */
.badge {
  display: inline-block; padding: 2px 8px; border-radius: 10px;
  font-size: 11px; font-weight: 600; white-space: nowrap;
}
.badge-решено     { background: #d1fae5; color: #065f46; }
.badge-не-решено  { background: #fee2e2; color: #991b1b; }
.badge-частично   { background: #fef3c7; color: #92400e; }
.badge-эскалация  { background: #ede9fe; color: #5b21b6; }
.badge-other      { background: #f3f4f6; color: #374151; }

/* Rating / Complexity widget */
.rating-widget { display: flex; gap: 2px; align-items: center; }
.rating-dot {
  width: 22px; height: 22px; border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 11px; font-weight: 700; cursor: default;
  border: 1.5px solid #e0e0e0; color: #bbb; background: #fafafa;
  transition: background .15s, border-color .15s, color .15s;
  user-select: none; flex-shrink: 0;
}
.rating-dot.filled-rating     { background: #e8f5e9; border-color: #66bb6a; color: #2e7d32; }
.rating-dot.filled-complexity { background: #fff3e0; border-color: #ffa726; color: #e65100; }
.rating-dot.clickable { cursor: pointer; }
.rating-dot.clickable:hover   { border-color: #1a73e8; color: #1a73e8; background: #e8f0fe; }

/* Problem summary truncation */
.summary-text {
  display: -webkit-box;
  -webkit-line-clamp: 3;
  -webkit-box-orient: vertical;
  overflow: hidden;
  cursor: pointer;
}
.summary-text.expanded {
  display: block;
  -webkit-line-clamp: unset;
}

.no-data { text-align: center; padding: 60px 20px; color: #aaa; font-size: 14px; }

/* ── Кнопка и панель управления столбцами ── */
.btn-cols {
  display: flex; align-items: center; gap: 5px;
  padding: 5px 11px; border-radius: 6px; border: 1px solid #d0d7e3;
  background: #fff; font-size: 12px; color: #444; cursor: pointer;
  transition: background .15s, border-color .15s;
}
.btn-cols:hover { background: #f0f7ff; border-color: #1a73e8; color: #1a73e8; }
.btn-cols svg { flex-shrink: 0; }

.col-panel {
  position: fixed; top: 50%; left: 50%; transform: translate(-50%,-50%);
  z-index: 10000; background: #fff; border-radius: 12px;
  box-shadow: 0 12px 48px rgba(0,0,0,.18); width: 280px;
  max-height: 80vh; display: flex; flex-direction: column;
  font-size: 13px;
}
.col-panel-head {
  display: flex; align-items: center; justify-content: space-between;
  padding: 14px 16px 10px; border-bottom: 1px solid #f0f0f0;
  font-weight: 600; color: #222; font-size: 14px;
}
.col-panel-head button {
  background: none; border: none; font-size: 18px; cursor: pointer;
  color: #aaa; line-height: 1; padding: 2px 4px; border-radius: 4px;
}
.col-panel-head button:hover { background: #f5f5f5; color: #333; }
.col-panel-hint {
  padding: 6px 16px 8px; font-size: 11px; color: #aaa; border-bottom: 1px solid #f0f0f0;
}
.col-panel-list {
  overflow-y: auto; padding: 6px 0; flex: 1;
}
.col-panel-item {
  display: flex; align-items: center; gap: 10px;
  padding: 7px 16px; cursor: grab; user-select: none;
  border-radius: 6px; margin: 1px 6px; transition: background .1s;
}
.col-panel-item:hover { background: #f5f8ff; }
.col-panel-item.dragging { opacity: .45; background: #e8f0fe; }
.col-panel-item.drag-over { border-top: 2px solid #1a73e8; }
.col-panel-item .drag-handle {
  color: #ccc; font-size: 14px; flex-shrink: 0; cursor: grab;
}
.col-panel-item input[type=checkbox] { cursor: pointer; accent-color: #1a73e8; width:15px; height:15px; flex-shrink:0; }
.col-panel-item label { cursor: pointer; flex: 1; }
.col-panel-foot {
  display: flex; gap: 8px; padding: 10px 16px 14px; border-top: 1px solid #f0f0f0;
}
.col-panel-foot button {
  flex: 1; padding: 7px; border-radius: 6px; border: 1px solid #d0d7e3;
  background: #fff; font-size: 12px; cursor: pointer; color: #444;
  transition: background .15s;
}
.col-panel-foot button.primary {
  background: #1a73e8; color: #fff; border-color: #1a73e8; font-weight: 600;
}
.col-panel-foot button.primary:hover { background: #1558b0; }
.col-panel-foot button:not(.primary):hover { background: #f5f5f5; }

.col-panel-backdrop {
  position: fixed; inset: 0; z-index: 9999;
  background: rgba(0,0,0,.25);
}

/* Manual rows */
tr[data-manual="1"] td:first-child { border-left: 2px solid #d0c8a0; }
tr[data-manual="1"] .editable[data-field="problem_summary"] {
  display: -webkit-box; -webkit-line-clamp: 3; -webkit-box-orient: vertical;
  overflow: hidden; cursor: text; white-space: normal;
}
tr[data-manual="1"] .editable[data-field="problem_summary"]:focus {
  display: block; -webkit-line-clamp: unset; overflow: visible;
}

.btn-green { background: #0f9d58; }
.btn-green:hover { background: #0b7a43; }

.del-btn {
  float: right; margin-left: 4px;
  background: none; border: none;
  color: #ccc; cursor: pointer;
  font-size: 15px; line-height: 1; padding: 0 2px;
}
.del-btn:hover { color: #e53e3e; }

/* Expand dialog */
.expand-btn {
  cursor: pointer; color: #ccc; font-size: 9px;
  margin-right: 4px; display: inline-block;
  transition: transform .15s, color .15s;
  user-select: none; vertical-align: middle;
}
.expand-btn:hover { color: #1a73e8; }
.expand-btn.open { transform: rotate(90deg); color: #1a73e8; }

tr.dialog-row td { padding: 0 !important; background: #f8f9fa; border-bottom: 2px solid #e0e0e0; }
.dialog-wrap { padding: 14px 20px; max-height: 420px; overflow-y: auto; display: flex; flex-direction: column; gap: 6px; }
.d-msg { display: flex; }
.d-msg.visitor { padding-left: 16px; }
.d-msg.agent   { padding-left: 0; }
.d-bubble-wrap { max-width: 65%; }
.d-name { font-size: 10px; color: #999; margin-bottom: 2px; }
.d-bubble {
  padding: 7px 11px; border-radius: 12px;
  font-size: 12px; line-height: 1.5; word-break: break-word; white-space: pre-wrap;
}
.d-bubble.visitor { background: #1a73e8; color: #fff; border-bottom-left-radius: 3px; }
.d-bubble.agent   { background: #fff; border: 1px solid #ddd; color: #333; border-bottom-left-radius: 3px; }
.d-time { font-size: 10px; color: #bbb; margin-top: 3px; }
.d-empty { color: #aaa; font-style: italic; font-size: 12px; padding: 8px 0; }
.spinner { display: inline-block; width: 16px; height: 16px; border: 2px solid #ddd; border-top-color: #1a73e8; border-radius: 50%; animation: spin .6s linear infinite; vertical-align: middle; margin-right: 6px; }
@keyframes spin { to { transform: rotate(360deg); } }

/* ── Пагинация ── */
.pagination {
  flex-shrink: 0;
  display: flex; align-items: center; justify-content: center; gap: 4px;
  padding: 8px; background: #fff; border-top: 1px solid #e8e8e8;
}
.pg-btn {
  min-width: 32px; height: 28px; padding: 0 8px;
  border: 1px solid #ddd; border-radius: 4px; background: #fff;
  font-size: 12px; cursor: pointer; color: #333;
  display: flex; align-items: center; justify-content: center;
}
.pg-btn:hover:not(:disabled) { background: #f0f7ff; border-color: #1a73e8; color: #1a73e8; }
.pg-btn.active { background: #1a73e8; color: #fff; border-color: #1a73e8; font-weight: 600; }
.pg-btn:disabled { opacity: .4; cursor: default; }
.pg-info { font-size: 12px; color: #888; padding: 0 8px; }
</style>
</head>
<body>

<div class="topbar">
  <h1>📋 Журнал обращений</h1>
  <div class="topbar-right">
    <div class="topbar-search">
      <span class="topbar-search-icon">🔍</span>
      <input type="text" id="filter_search" placeholder="Поиск по всем полям...">
    </div>
    %%ADMIN_LINK%%
    <div class="user-chip">
      <div class="user-avatar" id="user-avatar"></div>
      <span>%%USER%%</span>
    </div>
    <a href="/auth/logout" class="topbar-link">Выйти</a>
  </div>
</div>
<div class="filterbar">
  <div class="fg">
    <label>С</label>
    <input type="date" id="date_from">
  </div>
  <div class="fg">
    <label>По</label>
    <input type="date" id="date_to">
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Оператор</label>
    <select id="filter_operator"><option value="">Все</option></select>
  </div>
  <div class="fg">
    <label>Автор</label>
    <input type="text" id="filter_author" placeholder="введите имя" style="width:130px">
  </div>
  <div class="fg">
    <label>Логин</label>
    <input type="text" id="filter_login" placeholder="введите логин" style="width:130px">
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Тип</label>
    <select id="filter_stype">
      <option value="">Все</option>
      <option value="Клиент">Клиент</option>
      <option value="ПВЗ">ПВЗ</option>
      <option value="Сотрудник">Сотрудник</option>
    </select>
  </div>
  <div class="fg">
    <label>Канал</label>
    <select id="filter_channel"><option value="">Все</option></select>
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Категория</label>
    <select id="filter_category" onchange="updateSubcatFilter()">
      <option value="">Все</option>
    </select>
  </div>
  <div class="fg">
    <label>Подкатегория</label>
    <select id="filter_subcategory"><option value="">Все</option></select>
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Результат</label>
    <select id="filter_result">
      <option value="">Все</option>
      <option value="Решено">Решено</option>
      <option value="Не решено">Не решено</option>
      <option value="Эскалация">Эскалация</option>
    </select>
  </div>
  <div class="fg">
    <label>Отдел</label>
    <select id="filter_dept">
      <option value="">Все</option>
      <option value="Склад">Склад</option>
      <option value="ОКК">ОКК</option>
      <option value="ИТ">ИТ</option>
      <option value="Продукт">Продукт</option>
      <option value="Маркетинг">Маркетинг</option>
      <option value="Бухгалтерия">Бухгалтерия</option>
      <option value="Юрист">Юрист</option>
    </select>
  </div>
  <div class="filterbar-actions">
    <button class="btn" onclick="load()">Применить</button>
    <button class="btn btn-green" id="btn_add" onclick="addManualRow()">+ Строка</button>
    <button class="btn-cols" onclick="openColPanel()" title="Настроить столбцы">
      <svg width="14" height="14" viewBox="0 0 16 16" fill="currentColor">
        <rect x="1" y="1" width="4" height="14" rx="1"/><rect x="6" y="1" width="4" height="14" rx="1"/><rect x="11" y="1" width="4" height="14" rx="1"/>
      </svg>
      Столбцы
    </button>
    <button class="btn-cols" id="btn-export" onclick="exportExcel()" title="Выгрузить в Excel">
      <svg width="14" height="14" viewBox="0 0 16 16" fill="currentColor">
        <path d="M2 2h7l4 4v8a1 1 0 0 1-1 1H2a1 1 0 0 1-1-1V3a1 1 0 0 1 1-1z" fill="none" stroke="currentColor" stroke-width="1.2"/>
        <path d="M9 2v4h4M5 9l2 2.5L9 9M7 11.5V7" stroke="currentColor" stroke-width="1.2" fill="none" stroke-linecap="round"/>
      </svg>
      Excel
    </button>
    <span class="filterbar-count" id="count"></span>
  </div>
</div>

<div id="col-panel-backdrop" class="col-panel-backdrop" style="display:none" onclick="closeColPanel()"></div>
<div id="col-panel" class="col-panel" style="display:none">
  <div class="col-panel-head">
    <span>Настройка столбцов</span>
    <button onclick="closeColPanel()">×</button>
  </div>
  <div class="col-panel-hint">Перетащите для изменения порядка · Чекбокс — скрыть/показать</div>
  <div class="col-panel-list" id="col-panel-list"></div>
  <div class="col-panel-foot">
    <button onclick="resetColLayout()">Сбросить</button>
    <button class="primary" onclick="closeColPanel()">Готово</button>
  </div>
</div>

<div class="wrap">
  <table>
    <thead>
      <tr>
        <th data-col="date">Дата</th>
        <th data-col="time">Время</th>
        <th data-col="operator">Оператор</th>
        <th data-col="source_type">Тип автора</th>
        <th data-col="author">Автор</th>
        <th data-col="login">Логин</th>
        <th data-col="appeal_type">Тип</th>
        <th data-col="category">Категория</th>
        <th data-col="subcategory">Подкатегория</th>
        <th data-col="problem_summary">Причина обращения</th>
        <th data-col="result">Результат</th>
        <th data-col="comment">Комм. поддержки</th>
        <th data-col="comment_manager">Комм. менеджера</th>
        <th data-col="responsible_dept" class="col-dept">Отв. отдел</th>
        <th data-col="channel">Канал</th>
        <th data-col="chat_id">№ обращения</th>
        <th data-col="rating" style="text-align:center">Оценка</th>
        <th data-col="complexity" style="text-align:center">Сложность</th>
      </tr>
    </thead>
    <tbody id="tbody">
      <tr><td colspan="16" class="no-data"><span class="spinner"></span>Загрузка...</td></tr>
    </tbody>
  </table>
</div>

<div class="pagination" id="pagination"></div>

<script>
// Defaults: last 7 days
const today = new Date();
const week  = new Date(today); week.setDate(today.getDate() - 7);
document.getElementById('date_from').value = fmt(week);
document.getElementById('date_to').value   = fmt(today);

let currentPage = 1;

function fmt(d) {
  return d.toISOString().split('T')[0];  // YYYY-MM-DD для input[type=date]
}

// YYYY-MM-DD → ДД.ММ.ГГГГ для отображения
function fmtDate(s) {
  if (!s) return '';
  const p = s.split('-');
  if (p.length === 3) return p[2] + '.' + p[1] + '.' + p[0];
  return s;
}

// ДД.ММ.ГГГГ → YYYY-MM-DD для сохранения
function parseDate(s) {
  if (!s) return '';
  const p = s.split('.');
  if (p.length === 3 && p[2].length === 4) return p[2] + '-' + p[1] + '-' + p[0];
  return s; // уже YYYY-MM-DD или непонятный формат — оставляем как есть
}

function esc(s) {
  if (s === null || s === undefined) return '';
  return String(s)
    .replace(/&/g, '&amp;').replace(/</g, '&lt;')
    .replace(/>/g, '&gt;').replace(/"/g, '&quot;');
}

function badgeClass(result) {
  const r = (result || '').toLowerCase();
  if (r === 'решено')    return 'badge-решено';
  if (r === 'не решено') return 'badge-не-решено';
  if (r === 'частично')  return 'badge-частично';
  if (r === 'эскалация') return 'badge-эскалация';
  return r ? 'badge-other' : '';
}

const FILTERS_KEY = 'log_filters_v1';

function saveFilters() {
  try {
    localStorage.setItem(FILTERS_KEY, JSON.stringify({
      date_from:   document.getElementById('date_from').value,
      date_to:     document.getElementById('date_to').value,
      stype:       document.getElementById('filter_stype').value,
      category:    document.getElementById('filter_category').value,
      subcategory: document.getElementById('filter_subcategory').value,
      result:      document.getElementById('filter_result').value,
      dept:        document.getElementById('filter_dept').value,
      operator:    document.getElementById('filter_operator').value,
      channel:     document.getElementById('filter_channel').value,
      author:      document.getElementById('filter_author').value,
      login:       document.getElementById('filter_login').value,
      search:      document.getElementById('filter_search').value,
    }));
  } catch(_) {}
}

function restoreFilters() {
  try {
    const f = JSON.parse(localStorage.getItem(FILTERS_KEY) || '{}');
    if (!f || !Object.keys(f).length) return;
    const set = (id, val) => { if (val) { const el = document.getElementById(id); if (el) el.value = val; } };
    set('date_from', f.date_from);
    set('date_to',   f.date_to);
    set('filter_stype',       f.stype);
    set('filter_result',      f.result);
    set('filter_dept',        f.dept);
    set('filter_author',      f.author);
    set('filter_login',       f.login);
    set('filter_search',      f.search);
    if (f.category) {
      set('filter_category', f.category);
      updateSubcatFilter();
      set('filter_subcategory', f.subcategory);
    }
    // operator и channel — динамические, восстанавливаем через data-restore после загрузки опций
    if (f.operator) document.getElementById('filter_operator').dataset.restore = f.operator;
    if (f.channel)  document.getElementById('filter_channel').dataset.restore  = f.channel;
  } catch(_) {}
}

function initFilters() {
  const sel = document.getElementById('filter_category');
  sel.innerHTML = '<option value="">Все</option>' +
    Object.keys(CAT_MAP).map(c => `<option value="${esc(c)}">${esc(c)}</option>`).join('');
  restoreFilters();

  // Автопоиск по Автору, Логину и Поиску от 3 символов с debounce 400ms
  let _searchTimer = null;
  ['filter_author', 'filter_login', 'filter_search'].forEach(id => {
    const el = document.getElementById(id);
    if (!el) return;
    el.addEventListener('input', function() {
      clearTimeout(_searchTimer);
      const val = this.value;
      if (val.length === 0 || val.length >= 3) {
        _searchTimer = setTimeout(() => load(), 400);
      }
    });
  });
}

function updateSubcatFilter() {
  const cat = document.getElementById('filter_category').value;
  const sel = document.getElementById('filter_subcategory');
  const cur = sel.value;
  const subs = CAT_MAP[cat] || [];
  sel.innerHTML = '<option value="">Все</option>' +
    subs.map(s => `<option value="${esc(s)}"${s===cur?' selected':''}>${esc(s)}</option>`).join('');
}

function goPage(p) { currentPage = p; load(false); }

function buildFilterParams() {
  const params = new URLSearchParams();
  const v = id => (document.getElementById(id) || {}).value || '';
  params.set('date_from',  v('date_from'));
  params.set('date_to',    v('date_to'));
  params.set('operator',   v('filter_operator'));
  params.set('channel',    v('filter_channel'));
  params.set('stype',      v('filter_stype'));
  params.set('category',   v('filter_category'));
  params.set('subcategory',v('filter_subcategory'));
  params.set('result',     v('filter_result'));
  params.set('dept',       v('filter_dept'));
  params.set('author',     v('filter_author'));
  params.set('login',      v('filter_login'));
  params.set('search',     v('filter_search') || '');
  return params;
}

async function exportExcel() {
  const btn = document.getElementById('btn-export');
  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span> Формируем...';
  try {
    const params = buildFilterParams();
    const resp = await fetch('/api/log/export?' + params.toString());
    if (!resp.ok) {
      const err = await resp.json().catch(() => ({}));
      alert(err.error || 'Ошибка экспорта');
      return;
    }
    const blob = await resp.blob();
    const url  = URL.createObjectURL(blob);
    const a    = document.createElement('a');
    a.href     = url;
    const cd   = resp.headers.get('Content-Disposition') || '';
    const m    = cd.match(/filename\*?=(?:UTF-8'')?([^;]+)/i);
    a.download = m ? decodeURIComponent(m[1]) : 'обращения.xlsx';
    a.click();
    URL.revokeObjectURL(url);
  } catch(e) {
    alert('Ошибка: ' + e.message);
  } finally {
    btn.disabled = false;
    btn.innerHTML = '<svg width="14" height="14" viewBox="0 0 16 16" fill="currentColor"><path d="M2 2h7l4 4v8a1 1 0 0 1-1 1H2a1 1 0 0 1-1-1V3a1 1 0 0 1 1-1z" fill="none" stroke="currentColor" stroke-width="1.2"/><path d="M9 2v4h4M5 9l2 2.5L9 9M7 11.5V7" stroke="currentColor" stroke-width="1.2" fill="none" stroke-linecap="round"/></svg> Excel';
  }
}

async function load(resetPage = true) {
  if (resetPage) currentPage = 1;

  const tbody = document.getElementById('tbody');
  tbody.innerHTML = '<tr><td colspan="16" class="no-data"><span class="spinner"></span>Загрузка...</td></tr>';
  document.getElementById('count').textContent = '';
  document.getElementById('pagination').innerHTML = '';

  try {
    // dr(id) — берёт value или data-restore (для динамических селектов до загрузки опций)
    const dr = id => { const el = document.getElementById(id); return el ? (el.value || el.dataset.restore || '') : ''; };
    const gv = id => { const el = document.getElementById(id); return el ? el.value : ''; };

    const params = new URLSearchParams({
      date_from:   gv('date_from'),
      date_to:     gv('date_to'),
      operator:    dr('filter_operator'),
      channel:     dr('filter_channel'),
      stype:       gv('filter_stype'),
      category:    gv('filter_category'),
      subcategory: gv('filter_subcategory'),
      result:      gv('filter_result'),
      dept:        gv('filter_dept'),
      author:      gv('filter_author'),
      login:       gv('filter_login'),
      search:      gv('filter_search'),
      page:        currentPage,
    });

    saveFilters();

    const resp = await fetch('/api/log?' + params);
    if (!resp.ok) throw new Error('HTTP ' + resp.status);
    const data = await resp.json();

    // Обновляем списки операторов и каналов (только при первой загрузке / сбросе)
    if (resetPage) {
      const selOp = document.getElementById('filter_operator');
      const curOp = selOp.value || selOp.dataset.restore || '';
      delete selOp.dataset.restore;
      selOp.innerHTML = '<option value="">Все</option>' +
        data.operators.map(o => `<option value="${esc(o)}"${o===curOp?' selected':''}>${esc(o)}</option>`).join('');
      if (curOp) selOp.value = curOp;

      const selCh = document.getElementById('filter_channel');
      const curCh = selCh.value || selCh.dataset.restore || '';
      delete selCh.dataset.restore;
      selCh.innerHTML = '<option value="">Все</option>' +
        (data.channels||[]).map(c => `<option value="${esc(c)}"${c===curCh?' selected':''}>${esc(c)}</option>`).join('');
      if (curCh) selCh.value = curCh;
    }

    render(data.rows);
    highlightSearch((document.getElementById('filter_search') || {}).value || '');
    applyHiddenCols();
    applyColLayout();
    renderPagination(data.page, data.pages, data.total);
    document.getElementById('count').textContent =
      `${data.total} записей`;
  } catch (e) {
    tbody.innerHTML = `<tr><td colspan="16" class="no-data">Ошибка загрузки: ${e.message}</td></tr>`;
  }
}

function highlightSearch(term) {
  if (!term || term.length < 3) return;
  const tbody = document.getElementById('tbody');
  const lower = term.toLowerCase();

  function walkNode(node) {
    if (node.nodeType === 3) {
      const text = node.textContent;
      const ltext = text.toLowerCase();
      if (ltext.indexOf(lower) === -1) return;
      const frag = document.createDocumentFragment();
      let pos = 0, i;
      while ((i = ltext.indexOf(lower, pos)) !== -1) {
        if (i > pos) frag.appendChild(document.createTextNode(text.slice(pos, i)));
        const mark = document.createElement('mark');
        mark.className = 'hl';
        mark.textContent = text.slice(i, i + term.length);
        frag.appendChild(mark);
        pos = i + term.length;
      }
      if (pos < text.length) frag.appendChild(document.createTextNode(text.slice(pos)));
      node.parentNode.replaceChild(frag, node);
    } else if (node.nodeType === 1) {
      if (node.contentEditable === 'true' || node.tagName === 'SELECT' || node.tagName === 'INPUT') return;
      Array.from(node.childNodes).forEach(walkNode);
    }
  }

  tbody.querySelectorAll('td').forEach(td => walkNode(td));
}

function renderPagination(page, pages, total) {
  const el = document.getElementById('pagination');
  if (pages <= 1) { el.innerHTML = ''; return; }

  const btns = [];

  // Кнопка «Назад»
  btns.push(`<button class="pg-btn" onclick="goPage(${page-1})" ${page===1?'disabled':''}>‹</button>`);

  // Номера страниц (окно ±2 от текущей)
  const range = new Set([1, pages]);
  for (let i = Math.max(1, page-2); i <= Math.min(pages, page+2); i++) range.add(i);
  let prev = 0;
  for (const p of [...range].sort((a,b)=>a-b)) {
    if (prev && p - prev > 1) btns.push(`<span class="pg-info">…</span>`);
    btns.push(`<button class="pg-btn${p===page?' active':''}" onclick="goPage(${p})">${p}</button>`);
    prev = p;
  }

  // Кнопка «Вперёд»
  btns.push(`<button class="pg-btn" onclick="goPage(${page+1})" ${page===pages?'disabled':''}>›</button>`);

  el.innerHTML = btns.join('');
}

function applyHiddenCols() {
  const hidden = (PERMS.hidden || []);
  if (!hidden.length) return;
  const allThs = [...document.querySelectorAll('thead th')];
  hidden.forEach(function(col) {
    const th = document.querySelector('thead th[data-col="' + col + '"]');
    if (!th) return;
    const idx = allThs.indexOf(th) + 1;
    const s = document.createElement('style');
    s.textContent = 'thead th:nth-child(' + idx + '), tbody td:nth-child(' + idx + ') { display:none !important; }';
    document.head.appendChild(s);
  });
}

function canEdit(field) { return PERMS.editable.includes(field); }

function selCell(field, value) {
  if (canEdit(field)) {
    return `<div class="select-cell" data-field="${field}" data-value="${esc(value)}" onclick="openSelect(this)">${esc(value) || '<span style=color:#bbb>—</span>'}</div>`;
  }
  return `<span>${esc(value) || '<span style="color:#bbb">—</span>'}</span>`;
}

function editCell(field, value, placeholder) {
  if (canEdit(field)) {
    return `<div class="editable" contenteditable="true" data-field="${field}" data-orig="${esc(value)}" data-placeholder="${placeholder}">${esc(value)}</div>`;
  }
  return `<span>${esc(value)}</span>`;
}

function commentCell(field, value, author, placeholder) {
  const authorHtml = author ? `<span class="comment-author">${esc(author)}:</span>` : '';
  const isOwner = !author || author === CURRENT_USER;
  const isLead  = PERMS && PERMS.is_lead;

  if (!canEdit(field)) {
    // Нет прав на редактирование — только просмотр.
    // Скрытый .editable нужен чтобы collectFields/resolveAuthor видели значение
    // и не затирали comment + comment_author при сохранении других полей.
    return `<div class="comment-wrap">` +
      `<span class="editable" data-field="${field}" data-orig="${esc(value)}" data-orig-author="${esc(author)}" style="display:none">${esc(value)}</span>` +
      `${authorHtml}<span>${esc(value) || '<span style="color:#bbb">—</span>'}</span>` +
      `</div>`;
  }
  if (!value || isOwner || isLead) {
    // Пустое поле / свой комментарий / руководитель — редактируемо
    return `<div class="comment-wrap">${authorHtml}<div class="editable" contenteditable="true" data-field="${field}" data-orig="${esc(value)}" data-orig-author="${esc(author)}" data-placeholder="${placeholder}">${esc(value)}</div></div>`;
  }
  // Чужой комментарий, не руководитель: readonly + кнопка «+ Добавить»
  return `<div class="comment-wrap">${authorHtml}` +
    `<div class="comment-ro">${esc(value)}</div>` +
    `<span class="editable" data-field="${field}" data-orig="${esc(value)}" data-orig-author="${esc(author)}" style="display:none">${esc(value)}</span>` +
    `<button class="btn-append-comment" onclick="showAppend(this)">+ Добавить</button>` +
    `<div class="comment-append-area" style="display:none">` +
      `<div class="comment-append-input" contenteditable="true" data-placeholder="Ваш комментарий..."></div>` +
      `<div class="comment-append-actions">` +
        `<button class="btn-app-save" onclick="saveAppend(this,'${field}')">Сохранить</button>` +
        `<button class="btn-app-cancel" onclick="hideAppend(this)">Отмена</button>` +
      `</div>` +
    `</div>` +
  `</div>`;
}

function showAppend(btn) {
  btn.style.display = 'none';
  const area = btn.nextElementSibling; // .comment-append-area
  area.style.display = '';
  area.querySelector('.comment-append-input').focus();
}

function hideAppend(anyBtn) {
  const area    = anyBtn.closest('.comment-append-area');
  const showBtn = area.previousElementSibling; // .btn-append-comment
  area.querySelector('.comment-append-input').textContent = '';
  area.style.display = 'none';
  showBtn.style.display = '';
}

async function saveAppend(btn, field) {
  const area    = btn.closest('.comment-append-area');
  const input   = area.querySelector('.comment-append-input');
  const newText = input.textContent.trim();
  if (!newText) { hideAppend(btn); return; }

  const wrap   = area.closest('.comment-wrap');
  const row    = area.closest('tr');
  const rowKey = row.dataset.id;

  const hiddenSpan   = wrap.querySelector(`.editable[data-field="${field}"]`);
  const existingText = hiddenSpan ? hiddenSpan.textContent : '';
  const combined     = existingText + '\\n\\n' + CURRENT_USER + ': ' + newText;

  // Полный payload чтобы не затереть другие поля
  const allFields  = collectFields(row);
  allFields[field] = combined;

  // Автора оригинального комментария НЕ меняем
  const authorSpan = wrap.querySelector('.comment-author');
  const origAuthor = authorSpan ? authorSpan.textContent.replace(/:$/, '').trim() : '';
  if (field === 'comment')         allFields.comment_author         = origAuthor;
  if (field === 'comment_manager') allFields.comment_manager_author = origAuthor;

  // Сохраняем автора другого комментария
  const otherField     = field === 'comment' ? 'comment_manager' : 'comment';
  const otherAuthorKey = field === 'comment' ? 'comment_manager_author' : 'comment_author';
  const otherEl        = row.querySelector(`.editable[data-field="${otherField}"]`);
  if (otherEl) {
    const otherWrap = otherEl.closest && otherEl.closest('.comment-wrap');
    const otherSp   = otherWrap && otherWrap.querySelector('.comment-author');
    allFields[otherAuthorKey] = otherSp ? otherSp.textContent.replace(/:$/, '').trim() : '';
  }

  const url = rowKey.startsWith('m_') ? '/api/log/manual/' + rowKey.slice(2) : '/api/log/' + rowKey;
  btn.disabled = true;
  try {
    const r = await fetch(url, { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(allFields) });
    if (r.ok) {
      if (hiddenSpan) { hiddenSpan.textContent = combined; hiddenSpan.dataset.orig = combined; }
      const roDiv = wrap.querySelector('.comment-ro');
      if (roDiv) roDiv.textContent = combined;
      hideAppend(btn);
    }
  } catch(e) { console.error(e); }
  btn.disabled = false;
}

function deptCell(value) {
  const inner = value ? `<span class="dept-badge">${esc(value)}</span>` : '<span style="color:#bbb">—</span>';
  if (canEdit('responsible_dept')) {
    return `<div class="select-cell select-dept" data-field="responsible_dept" data-value="${esc(value)}" onclick="openSelect(this)">${inner}</div>`;
  }
  return inner;
}

function render(rows) {
  const tbody = document.getElementById('tbody');
  if (!rows.length) {
    tbody.innerHTML = '<tr><td colspan="16" class="no-data">Нет данных за выбранный период</td></tr>';
    return;
  }

  tbody.innerHTML = rows.map(r => {
    const rowKey  = r.row_key !== undefined ? r.row_key : String(r.chat_id);
    const isManual = !!r.is_manual;
    const chClass = r.channel === 'Чат' ? 'ch-chat' : r.channel === 'ЛС' ? 'ch-ls' : r.channel === 'Жалоба' ? 'ch-form' : '';
    const bc = badgeClass(r.result);

    // Результат: manual — всегда select; manager — select; support — badge
    const resultHtml = isManual || canEdit('result')
      ? `<div class="select-cell" data-field="result" data-value="${esc(r.result)}" onclick="openSelect(this)">${bc ? `<span class="badge ${bc}">${esc(r.result)}</span>` : (esc(r.result) || '<span style=color:#bbb>—</span>')}</div>`
      : (bc ? `<span class="badge ${bc}">${esc(r.result)}</span>` : (r.result ? esc(r.result) : '<span style="color:#bbb">—</span>'));

    const channelHtml = isManual
      ? `<div class="select-cell" data-field="channel" data-value="${esc(r.channel)}" onclick="openSelect(this)">${esc(r.channel) || '<span style=color:#bbb>—</span>'}</div>`
      : esc(r.channel);

    const dateHtml = isManual
      ? `<button class="del-btn" onclick="deleteManualRow(this,'${rowKey}')" title="Удалить строку">×</button><div class="editable" contenteditable="true" data-field="date" data-orig="${esc(r.date)}" data-placeholder="ДД.ММ.ГГГГ">${fmtDate(r.date)}</div>`
      : `<span class="expand-btn" onclick="toggleDialog(event,'${rowKey}')">▶</span>${fmtDate(r.date)}`;
    const timeHtml     = isManual ? `<div class="editable" contenteditable="true" data-field="time"     data-orig="${esc(r.time)}"     data-placeholder="ЧЧ:ММ">${esc(r.time)}</div>`     : esc(r.time);
    const operatorHtml = isManual ? `<div class="editable" contenteditable="true" data-field="operator" data-orig="${esc(r.operator)}" data-placeholder="Оператор">${esc(r.operator)}</div>` : esc(r.operator);
    const authorHtml   = isManual ? `<div class="editable" contenteditable="true" data-field="author"   data-orig="${esc(r.author)}"   data-placeholder="Автор">${esc(r.author)}</div>`     : esc(r.author);
    const appealHtml   = isManual ? `<div class="editable" contenteditable="true" data-field="appeal_type" data-orig="${esc(r.appeal_type)}" data-placeholder="Тип">${esc(r.appeal_type)}</div>` : esc(r.appeal_type);
    const summaryHtml  = isManual
      ? `<div class="editable" contenteditable="true" data-field="problem_summary" data-orig="${esc(r.problem_summary)}" data-placeholder="Суть обращения">${esc(r.problem_summary)}</div>`
      : `<div class="summary-text" onclick="this.classList.toggle('expanded')">${esc(r.problem_summary)}</div>`;

    const ratingVal     = parseInt(r.rating     || 0);
    const complexityVal = parseInt(r.complexity || 0);
    const canRate = canEdit('rating');

    return `<tr data-id="${rowKey}"${isManual ? ' data-manual="1"' : ''}>
      <td data-col="date" class="col-date">${dateHtml}</td>
      <td data-col="time" class="col-time">${timeHtml}</td>
      <td data-col="operator">${operatorHtml}</td>
      <td data-col="source_type">${selCell('source_type', r.source_type)}</td>
      <td data-col="author" class="col-author" title="${esc(r.author)}">${authorHtml}</td>
      <td data-col="login" class="col-login">${esc(extractLogin(r.author, r.source_type))}</td>
      <td data-col="appeal_type">${appealHtml}</td>
      <td data-col="category">${selCell('category', r.category)}</td>
      <td data-col="subcategory">${selCell('subcategory', r.subcategory)}</td>
      <td data-col="problem_summary" class="col-summary">${summaryHtml}</td>
      <td data-col="result">${resultHtml}</td>
      <td data-col="comment">${commentCell('comment', r.comment, r.comment_author, 'Добавить...')}</td>
      <td data-col="comment_manager">${commentCell('comment_manager', r.comment_manager, r.comment_manager_author, 'Добавить...')}</td>
      <td data-col="responsible_dept" class="col-dept">${isManual ? '<span style="color:#bbb">—</span>' : deptCell(r.responsible_dept || '')}</td>
      <td data-col="channel" class="col-channel ${chClass}">${channelHtml}</td>
      <td data-col="chat_id" class="col-login" style="color:#aaa">${isManual ? '' : esc(String(r.chat_id))}</td>
      <td data-col="rating" data-field="rating" data-value="${ratingVal}" style="text-align:center">${renderRatingWidget('rating', ratingVal, canRate)}</td>
      <td data-col="complexity" data-field="complexity" data-value="${complexityVal}" style="text-align:center">${renderRatingWidget('complexity', complexityVal, canRate)}</td>
    </tr>`;
  }).join('');

  tbody.querySelectorAll('.editable').forEach(el => {
    el.addEventListener('blur', onBlur);
  });
}

async function onBlur(e) {
  const el   = e.target;
  const row  = el.closest('tr');
  const orig = el.dataset.orig;
  const val  = el.textContent.trim();

  if (val === orig) return; // не изменилось

  // Очищаем <br> которые браузер оставляет при удалении текста,
  // чтобы CSS :empty::before (плейсхолдер) снова сработал
  if (val === '') el.innerHTML = '';

  el.dataset.orig = val;
  el.classList.add('saving');
  await saveRow(row, el);
  el.classList.remove('saving');

  // Обновляем отображение автора прямо в DOM после сохранения
  const field = el.dataset.field;
  if (field === 'comment' || field === 'comment_manager') {
    const wrap = el.closest('.comment-wrap');
    if (wrap) {
      let authorSpan = wrap.querySelector('.comment-author');
      if (val) {
        if (!authorSpan) {
          authorSpan = document.createElement('span');
          authorSpan.className = 'comment-author';
          wrap.insertBefore(authorSpan, wrap.firstChild);
        }
        const origAuthor = el.dataset.origAuthor || '';
        // origAuthor уже обновлён в resolveAuthor, читаем готовое значение
        authorSpan.textContent = (origAuthor || CURRENT_USER) + ':';
      } else if (authorSpan) {
        authorSpan.remove();
      }
    }
  }
}

// ── Логин из имени автора ──────────────────────────────────────────────────
function extractLogin(author, sourceType) {
  if (!author) return '';
  // ПВЗ и Сотрудник — логин равен автору
  if (sourceType === 'ПВЗ' || sourceType === 'Сотрудник') return author;

  // 1. После запятой
  const ci = author.indexOf(',');
  if (ci !== -1) return author.slice(ci + 1).trim();

  // 2. После слова "лог" (лог, логин и т.п.)
  const li = author.toLowerCase().indexOf('лог');
  if (li !== -1) {
    const after = author.slice(li + 3).replace(/^[\s:]+/, '').trim();
    if (after) return after;
  }

  // 3. Последнее слово
  const words = author.trim().split(/\s+/);
  if (words.length) return words[words.length - 1];

  // 4. Сам автор
  return author;
}

// ── Собрать все редактируемые поля строки ──────────────────────────────────
function collectFields(row) {
  const f = {};
  row.querySelectorAll('.editable').forEach(c => {
    let val = c.textContent.trim();
    if (c.dataset.field === 'date') val = parseDate(val); // ДД.ММ.ГГГГ → YYYY-MM-DD
    f[c.dataset.field] = val;
  });
  row.querySelectorAll('.select-cell[data-field]').forEach(c => {
    f[c.dataset.field] = c.dataset.value || '';
  });
  return f;
}

// ── Сохранить строку ────────────────────────────────────────────────────────
async function saveRow(row, feedbackEl) {
  const rowKey = row.dataset.id;
  const fields = collectFields(row);

  // Автор комментария:
  // - если текст не изменился → сохраняем прежнего автора из DOM
  // - если изменился и редактор = руководитель, а автор чужой → "Оригинал (ред. Руководитель)"
  // - иначе → текущий пользователь
  const commentEl    = row.querySelector('[data-field="comment"].editable');
  const commentMgrEl = row.querySelector('[data-field="comment_manager"].editable');
  function resolveAuthor(el, newVal, authorKey) {
    if (!el) return;
    if (newVal !== el.dataset.orig) {
      const origAuthor = el.dataset.origAuthor || '';
      if (PERMS.is_lead && origAuthor && origAuthor !== CURRENT_USER) {
        fields[authorKey] = origAuthor + ' (ред. ' + CURRENT_USER + ')';
      } else {
        fields[authorKey] = newVal ? CURRENT_USER : '';
      }
      el.dataset.origAuthor = fields[authorKey]; // обновляем на будущее
    } else {
      const sp = el.closest('.comment-wrap') && el.closest('.comment-wrap').querySelector('.comment-author');
      fields[authorKey] = sp ? sp.textContent.replace(/:$/, '').trim() : '';
    }
  }
  resolveAuthor(commentEl,    fields.comment,         'comment_author');
  resolveAuthor(commentMgrEl, fields.comment_manager, 'comment_manager_author');

  const url = (rowKey && rowKey.startsWith('m_'))
    ? '/api/log/manual/' + rowKey.slice(2)
    : '/api/log/' + rowKey;

  try {
    const resp = await fetch(url, {
      method:  'POST',
      headers: { 'Content-Type': 'application/json' },
      body:    JSON.stringify(fields),
    });
    if (resp.ok && feedbackEl) {
      feedbackEl.classList.add('saved');
      setTimeout(() => feedbackEl.classList.remove('saved'), 1200);
    }
  } catch (err) {
    console.error('Save error:', err);
  }
}

// ── Права доступа (подставляются сервером) ─────────────────────────────────
const PERMS        = %%PERMS%%;
const CURRENT_USER = "%%USER%%";

// Аватар — первая буква имени
(function() {
  const av = document.getElementById('user-avatar');
  if (av && CURRENT_USER) av.textContent = CURRENT_USER[0].toUpperCase();
})();

// ── Карта категорий / подкатегорий ─────────────────────────────────────────
const SOURCE_TYPES = ['Клиент', 'ПВЗ', 'Сотрудник', 'Жалоба', 'Заявка'];

const DEPTS = ['Склад','ОКК','ИТ','Продукт','Маркетинг','Бухгалтерия','Юрист'];

const CAT_MAP = {
  'Заказ и доставка':      ['Статус заказа / задержка','Заказ не найден / потерян','Статус не обновился','Объединение / формирование посылки','Возврат товара (логистика)'],
  'ПВЗ и выдача':          ['График / адрес ПВЗ','Заказ не выдается','Проблема при приеме товара','Не сканируется / нет стикера','Ручные операции / обход процесса','Стать раздающим'],
  'Оплата и финансы':      ['Возврат денег','Непонятные списания / долг','Платное хранение','Оргсбор — размер и расчёт'],
  'Технические проблемы':  ['Ошибка / сбой сайта','Некорректное отображение','Поиск / каталог','Мобильное приложение'],
  'Аккаунт и профиль':     ['Нет доступа / ошибка входа','Смена телефона / почты','Блокировка / ограничения','Управление рассылками'],
  'Качество товара':       ['Брак / дефект','Несоответствие описанию или фото','Пересорт / пришло не то','Повреждение при доставке','Нарушения — маркировка, сроки годности'],
  'Закупки и организаторы':['Вопрос по закупке','Статус закупки / оплаты','Жалоба на организатора'],
  'Пристрой':              ['Как работает пристрой','Проблемы / передача пристроя'],
  'Пожелания и инсайты':   ['Запрос нового бренда / поставщика','Запрос открытия ПВЗ в регионе','Запрос новой функции платформы','Коммерческое предложение / партнёрство','Сравнение с конкурентами (WB, Ozon)'],
  'Благодарность':         ['Благодарность'],
  'Не определено':         ['Служебное / внутреннее'],
};

// ── Рендер значения ячейки с нужным стилем ─────────────────────────────────
function renderCellInner(field, val) {
  if (!val) return '<span style="color:#bbb">—</span>';
  if (field === 'result') {
    const bc = badgeClass(val);
    return bc ? `<span class="badge ${bc}">${esc(val)}</span>` : esc(val);
  }
  if (field === 'responsible_dept') return `<span class="dept-badge">${esc(val)}</span>`;
  return esc(val);
}

// ── Рейтинг: рендер и клик ──────────────────────────────────────────────────
function renderRatingWidget(field, value, canEdit) {
  const filledClass = field === 'rating' ? 'filled-rating' : 'filled-complexity';
  const dots = [1, 2, 3].map(n => {
    const filled = n <= value ? filledClass : '';
    const click  = canEdit ? `onclick="setRating(this,${n})"` : '';
    return `<span class="rating-dot ${filled}${canEdit ? ' clickable' : ''}" data-n="${n}" ${click}>${n}</span>`;
  }).join('');
  return `<div class="rating-widget">${dots}</div>`;
}

async function setRating(dot, n) {
  const td     = dot.closest('td');
  const row    = dot.closest('tr');
  const field  = td.dataset.field;   // 'rating' или 'complexity'
  const rowKey = row.dataset.id;

  // Toggle: повторный клик на ту же цифру сбрасывает в 0
  const current = parseInt(td.dataset.value || '0');
  const newVal  = (current === n) ? 0 : n;
  td.dataset.value = newVal;
  td.innerHTML = renderRatingWidget(field, newVal, true);

  // Собираем ВСЕ поля строки чтобы не затереть другие данные при INSERT
  const fields = collectFields(row);

  // Читаем оба рейтинговых поля из TD (collectFields их не видит — нет .editable)
  const rTd = row.querySelector('td[data-field="rating"]');
  const cTd = row.querySelector('td[data-field="complexity"]');
  fields.rating     = rTd ? parseInt(rTd.dataset.value     || '0') : 0;
  fields.complexity = cTd ? parseInt(cTd.dataset.value     || '0') : 0;
  fields[field]     = newVal; // перезаписываем то, что только что изменили

  // Сохраняем авторов комментариев из DOM
  const commentEl    = row.querySelector('[data-field="comment"].editable');
  const commentMgrEl = row.querySelector('[data-field="comment_manager"].editable');
  function getAuthor(el) {
    if (!el) return '';
    const sp = el.closest('.comment-wrap') && el.closest('.comment-wrap').querySelector('.comment-author');
    return sp ? sp.textContent.replace(/:$/, '').trim() : '';
  }
  if (commentEl)    fields.comment_author         = getAuthor(commentEl);
  if (commentMgrEl) fields.comment_manager_author = getAuthor(commentMgrEl);

  const url = (rowKey && rowKey.startsWith('m_'))
    ? '/api/log/manual/' + rowKey.slice(2)
    : '/api/log/' + rowKey;
  try {
    await fetch(url, {
      method: 'POST',
      headers: {'Content-Type': 'application/json'},
      body: JSON.stringify(fields),
    });
    // вспышка сохранения
    td.querySelectorAll('.rating-dot').forEach(d => d.classList.add('saved'));
    setTimeout(() => td.querySelectorAll('.rating-dot').forEach(d => d.classList.remove('saved')), 800);
  } catch(e) { console.error('rating save error', e); }
}

// ── Открыть выпадающий список ──────────────────────────────────────────────
let _activeCellDd = null;
let _activeCellSrc = null;

function _cellDdOutside(e) {
  if (!_activeCellDd) return;
  if (_activeCellDd.contains(e.target)) return;
  if (e.target.closest && e.target.closest('.select-cell')) return;
  closeCellDd();
}

function closeCellDd() {
  if (_activeCellDd) {
    _activeCellDd.remove();
    _activeCellDd = null;
    _activeCellSrc = null;
    document.removeEventListener('click', _cellDdOutside);
  }
}

function openSelect(cell) {
  // Повторный клик по той же ячейке — закрыть
  if (_activeCellSrc === cell) { closeCellDd(); return; }
  closeCellDd();

  const field   = cell.dataset.field;
  const current = cell.dataset.value || '';
  const row     = cell.closest('tr');

  let options = [];
  if (field === 'source_type') {
    options = SOURCE_TYPES;
  } else if (field === 'category') {
    options = Object.keys(CAT_MAP);
  } else if (field === 'subcategory') {
    const catCell = row.querySelector('[data-field="category"]');
    const cat = catCell ? catCell.dataset.value : '';
    options = CAT_MAP[cat] || [];
  } else if (field === 'result') {
    options = ['Решено', 'Не решено', 'Эскалация'];
  } else if (field === 'channel') {
    options = ['Таблица', 'Заявка', 'ЛС', 'Email', 'Telegram', 'ВК', 'Жалоба'];
  } else if (field === 'responsible_dept') {
    options = DEPTS;
  }

  if (!options.length) return;

  const rect = cell.getBoundingClientRect();
  const dd = document.createElement('div');
  dd.className = 'cell-dd';
  dd.style.top  = (rect.bottom + window.scrollY + 2) + 'px';
  dd.style.left = (rect.left + window.scrollX) + 'px';

  options.forEach(v => {
    const opt = document.createElement('div');
    opt.className = 'cell-dd-opt' + (v === current ? ' active' : '');
    opt.textContent = v;
    opt.addEventListener('mousedown', async e => {
      e.preventDefault();
      e.stopPropagation();
      closeCellDd();
      cell.dataset.value = v;
      cell.innerHTML = renderCellInner(field, v);

      if (field === 'category') {
        const subCell = row.querySelector('[data-field="subcategory"]');
        if (subCell) {
          subCell.dataset.value = '';
          subCell.innerHTML = '<span style="color:#bbb">—</span>';
        }
      }

      await saveRow(row, cell);
    });
    dd.appendChild(opt);
  });

  document.body.appendChild(dd);
  _activeCellDd = dd;
  _activeCellSrc = cell;

  // Добавляем обработчик закрытия — после того как текущий click отработает
  setTimeout(() => {
    document.addEventListener('click', _cellDdOutside);
  }, 0);
}

// ── Управление столбцами: порядок и видимость ──────────────────────────────
const COL_ORDER_KEY  = 'log_col_order_v1';
const COL_HIDDEN_KEY = 'log_col_hidden_v1';

function getDefaultColOrder() {
  return [...document.querySelectorAll('thead th[data-col]')].map(th => th.dataset.col);
}
function loadColOrder()  { try { return JSON.parse(localStorage.getItem(COL_ORDER_KEY))  || getDefaultColOrder(); } catch(_){ return getDefaultColOrder(); } }
function saveColOrder(o) { localStorage.setItem(COL_ORDER_KEY, JSON.stringify(o)); }
function loadColHidden() { try { return JSON.parse(localStorage.getItem(COL_HIDDEN_KEY)) || []; } catch(_){ return []; } }
function saveColHidden(h){ localStorage.setItem(COL_HIDDEN_KEY, JSON.stringify(h)); }

function applyColLayout() {
  const order  = loadColOrder();
  const hidden = new Set(loadColHidden());
  const theadTr = document.querySelector('thead tr');
  if (!theadTr) return;

  // Переставляем <th> по порядку
  order.forEach(col => {
    const th = theadTr.querySelector('th[data-col="' + col + '"]');
    if (th) theadTr.appendChild(th);
  });

  // Переставляем <td> в каждой строке
  document.querySelectorAll('tbody tr').forEach(tr => {
    order.forEach(col => {
      const td = tr.querySelector('td[data-col="' + col + '"]');
      if (td) tr.appendChild(td);
    });
  });

  // Применяем видимость через <style>
  let style = document.getElementById('col-hidden-style');
  if (!style) { style = document.createElement('style'); style.id = 'col-hidden-style'; document.head.appendChild(style); }
  style.textContent = [...hidden].map(col =>
    'th[data-col="' + col + '"], td[data-col="' + col + '"] { display:none !important; }'
  ).join(' ');
}

// ── Панель настройки столбцов ──────────────────────────────────────────────
const COL_LABELS = {
  date:'Дата', time:'Время', operator:'Оператор', source_type:'Тип автора',
  author:'Автор', login:'Логин', appeal_type:'Тип', category:'Категория',
  subcategory:'Подкатегория', problem_summary:'Причина обращения',
  result:'Результат', comment:'Комм. поддержки', comment_manager:'Комм. менеджера',
  responsible_dept:'Отв. отдел', channel:'Канал', chat_id:'№ обращения',
  rating:'Оценка', complexity:'Сложность',
};

function openColPanel() {
  renderColPanelList();
  document.getElementById('col-panel-backdrop').style.display = '';
  document.getElementById('col-panel').style.display = '';
}
function closeColPanel() {
  document.getElementById('col-panel-backdrop').style.display = 'none';
  document.getElementById('col-panel').style.display = 'none';
}
function resetColLayout() {
  localStorage.removeItem(COL_ORDER_KEY);
  localStorage.removeItem(COL_HIDDEN_KEY);
  location.reload();
}

function renderColPanelList() {
  const order  = loadColOrder();
  const hidden = new Set(loadColHidden());
  const list   = document.getElementById('col-panel-list');
  list.innerHTML = '';

  order.forEach(col => {
    const li = document.createElement('div');
    li.className = 'col-panel-item';
    li.draggable = true;
    li.dataset.col = col;
    const checked = !hidden.has(col);
    li.innerHTML =
      '<span class="drag-handle">⠿</span>' +
      '<input type="checkbox" id="chk_' + col + '"' + (checked ? ' checked' : '') + '>' +
      '<label for="chk_' + col + '">' + (COL_LABELS[col] || col) + '</label>';

    li.querySelector('input').addEventListener('change', function() {
      const h = loadColHidden();
      if (this.checked) { const i = h.indexOf(col); if (i !== -1) h.splice(i,1); }
      else              { if (!h.includes(col)) h.push(col); }
      saveColHidden(h);
      applyColLayout();
    });

    // Drag-to-reorder
    li.addEventListener('dragstart', e => {
      e.dataTransfer.setData('text/plain', col);
      li.classList.add('dragging');
    });
    li.addEventListener('dragend', () => li.classList.remove('dragging'));
    li.addEventListener('dragover', e => { e.preventDefault(); li.classList.add('drag-over'); });
    li.addEventListener('dragleave', () => li.classList.remove('drag-over'));
    li.addEventListener('drop', e => {
      e.preventDefault();
      li.classList.remove('drag-over');
      const fromCol = e.dataTransfer.getData('text/plain');
      if (fromCol === col) return;
      const ord = loadColOrder();
      const fi  = ord.indexOf(fromCol);
      const ti  = ord.indexOf(col);
      if (fi === -1 || ti === -1) return;
      ord.splice(fi, 1);
      ord.splice(ti, 0, fromCol);
      saveColOrder(ord);
      applyColLayout();
      renderColPanelList();
    });

    list.appendChild(li);
  });
}

// ── Ресайз столбцов ────────────────────────────────────────────────────────
const COL_WIDTHS_KEY = 'log_col_widths_v2';

// Дата, Время, Оператор, Тип автора, Автор, Логин, Тип, Категория, Подкатегория, Причина обращения, Результат, Оценка, Сложность, Комм.поддержки, Комм.менеджера, Отв.отдел, Канал, № обращения
const DEFAULT_COL_WIDTHS = [90, 55, 90, 90, 100, 110, 90, 120, 140, 340, 90, 70, 70, 140, 140, 110, 55, 90];

function initResizableColumns() {
  const ths = [...document.querySelectorAll('thead th')];
  ths.forEach((th, i) => {
    // ── Ресайз ──
    const handle = document.createElement('div');
    handle.className = 'col-resize-handle';
    th.appendChild(handle);

    let startX, startW;
    handle.addEventListener('mousedown', e => {
      startX = e.pageX;
      startW = th.offsetWidth;
      handle.classList.add('dragging');
      document.body.style.cursor = 'col-resize';
      document.body.style.userSelect = 'none';
      const onMove = e => {
        const w = Math.max(30, startW + e.pageX - startX);
        th.style.width = th.style.minWidth = th.style.maxWidth = w + 'px';
      };
      const onUp = () => {
        handle.classList.remove('dragging');
        document.body.style.cursor = '';
        document.body.style.userSelect = '';
        document.removeEventListener('mousemove', onMove);
        document.removeEventListener('mouseup', onUp);
        saveColWidths();
      };
      document.addEventListener('mousemove', onMove);
      document.addEventListener('mouseup', onUp);
      e.preventDefault();
    });

    // ── Drag-to-reorder ──
    if (!th.dataset.col) return;
    th.draggable = true;
    th.addEventListener('dragstart', e => {
      e.dataTransfer.setData('text/plain', th.dataset.col);
      e.dataTransfer.effectAllowed = 'move';
      setTimeout(() => th.classList.add('th-dragging'), 0);
    });
    th.addEventListener('dragend', () => {
      th.classList.remove('th-dragging');
      document.querySelectorAll('thead th').forEach(t => t.classList.remove('th-drag-over'));
    });
    th.addEventListener('dragover', e => {
      if (!e.dataTransfer.types.includes('text/plain')) return;
      e.preventDefault();
      e.dataTransfer.dropEffect = 'move';
      document.querySelectorAll('thead th').forEach(t => t.classList.remove('th-drag-over'));
      if (th.dataset.col !== e.dataTransfer.getData('text/plain')) th.classList.add('th-drag-over');
    });
    th.addEventListener('dragleave', () => th.classList.remove('th-drag-over'));
    th.addEventListener('drop', e => {
      e.preventDefault();
      th.classList.remove('th-drag-over');
      const fromCol = e.dataTransfer.getData('text/plain');
      const toCol   = th.dataset.col;
      if (!fromCol || fromCol === toCol) return;
      const ord = loadColOrder();
      const fi  = ord.indexOf(fromCol);
      const ti  = ord.indexOf(toCol);
      if (fi === -1 || ti === -1) return;
      ord.splice(fi, 1);
      ord.splice(ti, 0, fromCol);
      saveColOrder(ord);
      applyColLayout();
    });
  });

  restoreColWidths();
}

function saveColWidths() {
  const widths = [...document.querySelectorAll('thead th')]
    .map(th => th.style.width || '');
  localStorage.setItem(COL_WIDTHS_KEY, JSON.stringify(widths));
}

function restoreColWidths() {
  try {
    const saved = JSON.parse(localStorage.getItem(COL_WIDTHS_KEY) || '[]');
    const ths = [...document.querySelectorAll('thead th')];
    ths.forEach((th, i) => {
      const w = saved[i] || (DEFAULT_COL_WIDTHS[i] ? DEFAULT_COL_WIDTHS[i] + 'px' : null);
      if (w) th.style.width = th.style.minWidth = th.style.maxWidth = w;
    });
  } catch (_) {}
}

// ── Диалог (разворачивание строки) ──────────────────────────────────────────
const _dialogCache = {};

async function toggleDialog(e, rowKey) {
  e.stopPropagation();
  const btn = e.currentTarget;
  const tr  = btn.closest('tr');
  const next = tr.nextElementSibling;

  // Свернуть если уже открыт
  if (next && next.dataset.dialogFor === rowKey) {
    next.remove();
    btn.classList.remove('open');
    return;
  }

  btn.textContent = '…';
  btn.classList.remove('open');

  try {
    let data = _dialogCache[rowKey];
    if (!data) {
      const resp = await fetch('/api/log/dialog/' + rowKey);
      data = await resp.json();
      _dialogCache[rowKey] = data;
    }
    const expandTr = document.createElement('tr');
    expandTr.className = 'dialog-row';
    expandTr.dataset.dialogFor = rowKey;
    expandTr.innerHTML = `<td colspan="16"><div class="dialog-wrap">${renderDialog(data)}</div></td>`;
    tr.after(expandTr);
    btn.textContent = '▶';
    btn.classList.add('open');
  } catch(err) {
    btn.textContent = '▶';
    console.error('Dialog load error:', err);
  }
}

function fmtMsgTime(ts) {
  if (!ts) return '';
  const d = new Date(ts > 1e10 ? ts : ts * 1000);
  return d.toLocaleTimeString('ru-RU', { hour: '2-digit', minute: '2-digit' });
}

function renderDialog(data) {
  const msgs = data.messages || [];
  if (!msgs.length) return `<div class="d-empty">Сообщения недоступны</div>`;
  return msgs.map(m => {
    const isV  = m.type === 'visitor';
    const cls  = isV ? 'visitor' : 'agent';
    const name = isV ? esc(data.visitor || 'Клиент') : esc(data.operator || 'Оператор');
    const text = esc((m.message || '').trim());
    const ts   = m.timestamp || m.ts || m.time || 0;
    const time = fmtMsgTime(ts);
    return `<div class="d-msg ${cls}">
      <div class="d-bubble-wrap">
        <div class="d-name">${name}</div>
        <div class="d-bubble ${cls}">${text}</div>
        ${time ? `<div class="d-time">${time}</div>` : ''}
      </div>
    </div>`;
  }).join('');
}

// ── Удаление ручной строки ──────────────────────────────────────────────────
async function deleteManualRow(btn, rowKey) {
  if (!confirm('Удалить строку?')) return;
  const id = rowKey.slice(2);
  try {
    const resp = await fetch('/api/log/manual/' + id, { method: 'DELETE' });
    if (resp.ok) btn.closest('tr').remove();
    else alert('Ошибка удаления');
  } catch(e) {
    alert('Ошибка: ' + e.message);
  }
}

// ── Добавление ручной строки ────────────────────────────────────────────────
async function addManualRow() {
  const btn = document.getElementById('btn_add');
  btn.disabled = true;
  try {
    const resp = await fetch('/api/log/manual', { method: 'POST' });
    const data = await resp.json();
    if (!data.ok) throw new Error('Ошибка создания');
    prependManualRow(data.id, data.date, data.time);
  } catch(e) {
    alert('Ошибка: ' + e.message);
  } finally {
    btn.disabled = false;
  }
}

function prependManualRow(id, dateStr, timeStr) {
  const rowKey  = 'm_' + id;
  const today   = dateStr || fmt(new Date());
  const nowTime = timeStr || '';
  const tr = document.createElement('tr');
  tr.dataset.id     = rowKey;
  tr.dataset.manual = '1';
  tr.innerHTML = `
    <td class="col-date"><button class="del-btn" onclick="deleteManualRow(this,'${rowKey}')" title="Удалить строку">×</button><div class="editable" contenteditable="true" data-field="date" data-orig="${esc(today)}" data-placeholder="ДД.ММ.ГГГГ">${fmtDate(today)}</div></td>
    <td class="col-time"><div class="editable" contenteditable="true" data-field="time" data-orig="${esc(nowTime)}" data-placeholder="ЧЧ:ММ">${esc(nowTime)}</div></td>
    <td><div class="editable" contenteditable="true" data-field="operator" data-orig="" data-placeholder="Оператор"></div></td>
    <td><div class="select-cell" data-field="source_type" data-value="" onclick="openSelect(this)"><span style="color:#bbb">—</span></div></td>
    <td class="col-author"><div class="editable" contenteditable="true" data-field="author" data-orig="" data-placeholder="Автор"></div></td>
    <td class="col-login"></td>
    <td><div class="editable" contenteditable="true" data-field="appeal_type" data-orig="" data-placeholder="Тип"></div></td>
    <td><div class="select-cell" data-field="category"    data-value="" onclick="openSelect(this)"><span style="color:#bbb">—</span></div></td>
    <td><div class="select-cell" data-field="subcategory" data-value="" onclick="openSelect(this)"><span style="color:#bbb">—</span></div></td>
    <td><div class="editable" contenteditable="true" data-field="problem_summary" data-orig="" data-placeholder="Суть обращения"></div></td>
    <td><div class="select-cell" data-field="result"  data-value="" onclick="openSelect(this)"><span style="color:#bbb">—</span></div></td>
    <td><div class="editable" contenteditable="true" data-field="comment" data-orig="" data-placeholder="Добавить..."></div></td>
    <td><div class="editable" contenteditable="true" data-field="comment_manager" data-orig="" data-placeholder="Добавить..."></div></td>
    <td><div class="select-cell select-dept" data-field="responsible_dept" data-value="" onclick="openSelect(this)"><span style="color:#bbb">—</span></div></td>
    <td class="col-channel"><div class="select-cell" data-field="channel" data-value="" onclick="openSelect(this)"><span style="color:#bbb">—</span></div></td>
    <td></td>
  `;
  tr.querySelectorAll('.editable').forEach(el => el.addEventListener('blur', onBlur));
  const tbody = document.getElementById('tbody');
  if (tbody.querySelector('.no-data')) tbody.innerHTML = '';
  tbody.insertBefore(tr, tbody.firstChild);
  tr.querySelector('.editable').focus();
}

try { initFilters(); } catch(e) { console.error('initFilters error:', e); }
try { initResizableColumns(); } catch(e) { console.error('initResizableColumns error:', e); }
load();
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# Day-tracker — /day-tracker  (аналог /log + колонка «Ответственный отдел»)
# ---------------------------------------------------------------------------

class DayTrackerEdit(BaseModel):
    organizer:                Optional[str] = ""
    responsible:              Optional[str] = ""
    result:                   Optional[str] = ""
    comment:                  Optional[str] = ""
    comment_author:           Optional[str] = ""
    comment_manager:          Optional[str] = ""
    comment_manager_author:   Optional[str] = ""
    source_type:              Optional[str] = ""
    category:                 Optional[str] = ""
    subcategory:              Optional[str] = ""
    responsible_dept:         Optional[str] = ""


@router.get("/api/day-tracker")
def api_day_tracker(
    date_from:   str = Query(default=None),
    date_to:     str = Query(default=None),
    operator:    str = Query(default=""),
    channel:     str = Query(default=""),
    stype:       str = Query(default=""),
    category:    str = Query(default=""),
    subcategory: str = Query(default=""),
    result:      str = Query(default=""),
    dept:        str = Query(default=""),
    author:      str = Query(default=""),
    login:       str = Query(default=""),
    search:      str = Query(default=""),
    page:        int = Query(default=1, ge=1),
):
    df = date_from or str(date.today() - timedelta(days=7))
    dt = date_to   or str(date.today())

    def q(s): return s.replace("'", "\\'")

    where = f"toDate(d.event_timestamp) BETWEEN '{df}' AND '{dt}'"
    if operator:
        where += f" AND d.operator_name = '{q(operator)}'"
    if channel == "Чат":
        where += " AND d.source = 'jivo'"
    elif channel == "ЛС":
        where += " AND d.source = 'site_pm'"
    elif channel == "Жалоба":
        where += " AND d.source = 'claim'"
    if stype:
        where += f" AND if(e.source_type!='', e.source_type, ifNull(a.source_type,'')) = '{q(stype)}'"
    if category:
        where += f" AND if(e.category!='', e.category, ifNull(a.category,'')) = '{q(category)}'"
    if subcategory:
        where += f" AND if(e.subcategory!='', e.subcategory, ifNull(a.subcategory,'')) = '{q(subcategory)}'"
    if result:
        where += f" AND if(e.result!='' AND e.result IS NOT NULL, e.result, ifNull(a.resolution_status,'')) = '{q(result)}'"
    if dept:
        where += f" AND ifNull(t.responsible_dept, '') = '{q(dept)}'"
    if author:
        where += f" AND ifNull(d.visitor_name, '') ILIKE '%{q(author)}%'"
    if login and len(login) >= 3:
        where += f" AND ifNull(d.visitor_name, '') ILIKE '%{q(login)}%'"
    if search and len(search) >= 3:
        sq = q(search)
        where += (
            f" AND ("
            f"ifNull(d.visitor_name,'') ILIKE '%{sq}%'"
            f" OR ifNull(d.operator_name,'') ILIKE '%{sq}%'"
            f" OR ifNull(a.contact_reason,'') ILIKE '%{sq}%'"
            f" OR if(e.category!='',e.category,ifNull(a.category,'')) ILIKE '%{sq}%'"
            f" OR if(e.subcategory!='',e.subcategory,ifNull(a.subcategory,'')) ILIKE '%{sq}%'"
            f" OR ifNull(a.user_problem_summary,'') ILIKE '%{sq}%'"
            f" OR if(e.result!='' AND e.result IS NOT NULL,e.result,ifNull(a.resolution_status,'')) ILIKE '%{sq}%'"
            f" OR ifNull(e.comment,'') ILIKE '%{sq}%'"
            f" OR ifNull(e.comment_manager,'') ILIKE '%{sq}%'"
            f" OR ifNull(t.responsible_dept,'') ILIKE '%{sq}%'"
            f" OR ifNull(d.page_url,'') ILIKE '%{sq}%'"
            f")"
        )

    offset = (page - 1) * PER_PAGE

    total_rows = ch_query(f"""
        SELECT count() AS total
        FROM dialogs d
        JOIN (
            SELECT chat_id, max(received_at) AS ts
            FROM raw_dialogs
            WHERE event_name = 'chat_finished'
            GROUP BY chat_id
        ) r ON d.chat_id = r.chat_id
        LEFT JOIN (SELECT * FROM support_log_edits FINAL) e ON d.chat_id = e.chat_id
        LEFT JOIN dialog_analysis a ON d.chat_id = a.chat_id AND d.source = a.source
        LEFT JOIN (SELECT * FROM day_tracker_edits FINAL) t ON d.chat_id = t.chat_id
        WHERE {where}
        FORMAT JSONEachRow
    """)
    total = int((total_rows[0] if total_rows else {}).get("total", 0))

    rows = ch_query(f"""
        SELECT
            d.chat_id                                                              AS chat_id,
            toString(toDate(r.ts))                                                 AS date,
            substring(toString(r.ts), 12, 5)                                       AS time,
            ifNull(d.operator_name, '')                                            AS operator,
            ifNull(d.visitor_name, '')                                             AS author,
            toString(ifNull(d.visitor_id, 0))                                      AS login,
            ifNull(a.contact_reason, '')                                           AS appeal_type,
            if(e.source_type  != '', e.source_type,  ifNull(a.source_type, ''))   AS source_type,
            if(e.category     != '', e.category,     ifNull(a.category, ''))      AS category,
            if(e.subcategory  != '', e.subcategory,  ifNull(a.subcategory, ''))   AS subcategory,
            ifNull(a.user_problem_summary, '')                                     AS problem_summary,
            if(e.result IS NOT NULL AND e.result != '',
               e.result, ifNull(a.resolution_status, ''))                          AS result,
            ifNull(t.responsible_dept, '')                                         AS responsible_dept,
            ifNull(e.comment, '')                                                  AS comment,
            ifNull(e.comment_author, '')                                           AS comment_author,
            ifNull(e.comment_manager, '')                                          AS comment_manager,
            ifNull(e.comment_manager_author, '')                                   AS comment_manager_author,
            multiIf(d.source='jivo','Чат',d.source='site_pm','ЛС',d.source='claim','Жалоба',d.source) AS channel
        FROM dialogs d
        JOIN (
            SELECT chat_id, max(received_at) AS ts
            FROM raw_dialogs
            WHERE event_name = 'chat_finished'
            GROUP BY chat_id
        ) r ON d.chat_id = r.chat_id
        LEFT JOIN dialog_analysis a ON d.chat_id = a.chat_id AND d.source = a.source
        LEFT JOIN (SELECT * FROM support_log_edits FINAL) e ON d.chat_id = e.chat_id
        LEFT JOIN (SELECT * FROM day_tracker_edits FINAL) t ON d.chat_id = t.chat_id
        WHERE {where}
        ORDER BY r.ts DESC
        LIMIT {PER_PAGE} OFFSET {offset}
        FORMAT JSONEachRow
    """)

    operators = ch_query("""
        SELECT DISTINCT operator_name
        FROM dialogs
        WHERE operator_name != '' AND operator_name IS NOT NULL
        ORDER BY operator_name
        FORMAT JSONEachRow
    """)

    ch_sources = ch_query("""
        SELECT multiIf(source='jivo','Чат',source='site_pm','ЛС',source='claim','Жалоба',source) AS channel
        FROM dialogs
        GROUP BY source
        FORMAT JSONEachRow
    """)
    manual_channels = ch_query("""
        SELECT DISTINCT channel FROM manual_log_entries
        WHERE channel != '' AND channel IS NOT NULL
        ORDER BY channel
        FORMAT JSONEachRow
    """)
    channels = sorted({r["channel"] for r in ch_sources + manual_channels if r.get("channel")})

    return JSONResponse({
        "rows":      rows,
        "operators": [r["operator_name"] for r in operators],
        "channels":  channels,
        "total":     total,
        "page":      page,
        "per_page":  PER_PAGE,
        "pages":     max(1, -(-total // PER_PAGE)),
    })


@router.post("/api/day-tracker/{chat_id}")
def api_day_tracker_edit(chat_id: int, payload: DayTrackerEdit):
    # Общие правки → support_log_edits (те же данные что и в /log)
    log_row = json.dumps({
        "chat_id":                  chat_id,
        "organizer":                payload.organizer               or "",
        "responsible":              payload.responsible             or "",
        "result":                   payload.result                  or "",
        "comment":                  payload.comment                 or "",
        "comment_author":           payload.comment_author          or "",
        "comment_manager":          payload.comment_manager         or "",
        "comment_manager_author":   payload.comment_manager_author  or "",
        "source_type":              payload.source_type             or "",
        "category":                 payload.category                or "",
        "subcategory":              payload.subcategory             or "",
    }, ensure_ascii=False)
    ch_execute(
        "INSERT INTO support_log_edits "
        "(chat_id, organizer, responsible, result, comment, comment_author, "
        "comment_manager, comment_manager_author, source_type, category, subcategory) "
        "FORMAT JSONEachRow",
        data=log_row.encode("utf-8"),
    )

    # Ответственный отдел → day_tracker_edits
    dept_row = json.dumps({
        "chat_id":          chat_id,
        "responsible_dept": payload.responsible_dept or "",
    }, ensure_ascii=False)
    ch_execute(
        "INSERT INTO day_tracker_edits (chat_id, responsible_dept) FORMAT JSONEachRow",
        data=dept_row.encode("utf-8"),
    )

    return JSONResponse({"ok": True})


@router.get("/day-tracker", response_class=HTMLResponse)
def day_tracker_page(user: dict = Depends(require_user)):
    perms = get_permissions(user) or {"editable": [], "hidden": [], "role": "support", "is_manager": False}
    admin_link = (
        '<a href="/admin/perms" class="topbar-admin">⚙ Права</a>'
        if perms.get("is_lead") else ""
    )
    return (
        _DAY_HTML
        .replace("%%USER%%",       user.get("name") or user.get("username", ""))
        .replace("%%PERMS%%",      json.dumps(perms, ensure_ascii=False))
        .replace("%%ADMIN_LINK%%", admin_link)
    )


_DAY_HTML = """<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Day Tracker</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
html, body {
  height: 100%; overflow: hidden;
  font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
  font-size: 13px; background: #f0f2f5; color: #222;
  display: flex; flex-direction: column;
}

/* ── Topbar ── */
.topbar {
  flex-shrink: 0; background: #1a73e8;
  padding: 0 16px; height: 48px;
  display: flex; align-items: center; gap: 10px;
  box-shadow: 0 2px 6px rgba(0,0,0,.18); z-index: 10;
}
.topbar h1 {
  font-size: 15px; font-weight: 600; color: #fff;
  white-space: nowrap; letter-spacing: .2px;
}
.topbar-right { margin-left: auto; display: flex; align-items: center; gap: 4px; }
.topbar-search { position: relative; display: flex; align-items: center; }
.topbar-search input {
  border: none; border-radius: 20px;
  padding: 5px 12px 5px 30px;
  font-size: 12px; background: rgba(255,255,255,.18); color: #fff;
  width: 200px; outline: none; transition: background .2s, width .2s;
}
.topbar-search input::placeholder { color: rgba(255,255,255,.6); }
.topbar-search input:focus { background: rgba(255,255,255,.28); width: 240px; }
.topbar-search-icon {
  position: absolute; left: 9px; font-size: 12px;
  color: rgba(255,255,255,.7); pointer-events: none;
}
.user-chip {
  display: flex; align-items: center; gap: 6px;
  background: rgba(255,255,255,.18); border-radius: 20px;
  padding: 3px 10px 3px 4px; font-size: 12px; color: #fff; white-space: nowrap;
}
.user-avatar {
  width: 26px; height: 26px; border-radius: 50%;
  background: rgba(255,255,255,.35);
  display: flex; align-items: center; justify-content: center;
  font-size: 11px; font-weight: 700; color: #fff;
}
.topbar-link {
  font-size: 12px; color: rgba(255,255,255,.85); text-decoration: none;
  white-space: nowrap; padding: 5px 10px; border-radius: 5px; transition: background .15s;
}
.topbar-link:hover { background: rgba(255,255,255,.18); color: #fff; }
.topbar-admin {
  font-size: 12px; color: #ffd54f; text-decoration: none; font-weight: 500;
  white-space: nowrap; padding: 5px 10px; border-radius: 5px; transition: background .15s;
}
.topbar-admin:hover { background: rgba(255,255,255,.15); }
/* ── Filterbar ── */
.filterbar {
  flex-shrink: 0; background: #f8f9fa; border-bottom: 1px solid #e0e0e0;
  padding: 7px 16px; display: flex; align-items: center; flex-wrap: wrap; gap: 6px;
}
.fg { display: flex; align-items: center; gap: 4px; }
.fg label {
  font-size: 10.5px; color: #888; white-space: nowrap;
  font-weight: 600; text-transform: uppercase; letter-spacing: .4px;
}
.filterbar input[type=date], .filterbar input[type=text] {
  border: 1px solid #d5d5d5; border-radius: 6px; padding: 4px 8px;
  font-size: 12px; background: #fff; color: #333;
  height: 30px; box-sizing: border-box; transition: border-color .15s, box-shadow .15s;
}
.filterbar select {
  border: 1px solid #d5d5d5; border-radius: 6px; padding: 4px 26px 4px 9px;
  font-size: 12px; background: #fff; color: #333;
  height: 30px; box-sizing: border-box; transition: border-color .15s, box-shadow .15s;
  appearance: none; -webkit-appearance: none;
  background-image: url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6' viewBox='0 0 10 6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='%23999'/%3E%3C/svg%3E");
  background-repeat: no-repeat; background-position: right 8px center;
  cursor: pointer;
}
.filterbar input[type=date]:focus, .filterbar input[type=text]:focus, .filterbar select:focus {
  outline: none; border-color: #1a73e8; box-shadow: 0 0 0 2px rgba(26,115,232,.12);
}
.filterbar select:hover { border-color: #b0b0b0; }
.filterbar select option { padding: 6px; }
.filter-sep { width: 1px; height: 22px; background: #d8d8d8; margin: 0 2px; }
.filterbar-actions { display: flex; align-items: center; gap: 8px; margin-left: auto; }
.filterbar-count { font-size: 12px; color: #999; white-space: nowrap; padding-left: 4px; border-left: 1px solid #e0e0e0; }
.count { font-size: 12px; color: #888; white-space: nowrap; }
.btn {
  background: #1a73e8; color: #fff; border: none; border-radius: 6px;
  padding: 0 14px; height: 30px; cursor: pointer; font-size: 12px; font-weight: 500;
  white-space: nowrap; transition: background .15s;
}
.btn:hover { background: #1557b0; }

/* ── Table wrapper ── */
.wrap {
  flex: 1;
  overflow: auto;
  padding: 0 0 10px 0;
}

table {
  border-collapse: collapse;
  width: max-content; min-width: 100%;
  background: #fff;
  box-shadow: 0 1px 4px rgba(0,0,0,.08);
}

thead th {
  background: #f1f3f4;
  border-bottom: 2px solid #ddd;
  border-right: 1px solid #e0e0e0;
  padding: 8px 10px;
  text-align: left;
  font-weight: 600;
  font-size: 12px;
  color: #444;
  white-space: nowrap;
  position: sticky;
  top: 0;
  z-index: 10;
}
thead th:last-child { border-right: none; }

/* ── Ответственный отдел — выделенная колонка ── */
thead th.col-dept { background: #eef2ff; color: #3730a3; }
tbody td.col-dept { background: #f8f9ff; }

tbody tr { border-bottom: 1px solid #f0f0f0; }
tbody tr:last-child { border-bottom: none; }
tbody tr:hover { background: #f8fbff; }
tbody tr:hover td.col-dept { background: #eef2ff; }

tbody td {
  padding: 6px 10px;
  border-right: 1px solid #f0f0f0;
  vertical-align: top;
}
tbody td:last-child { border-right: none; }

.col-date    { white-space: nowrap; }
.col-time    { white-space: nowrap; color: #888; }
.col-channel { white-space: nowrap; font-weight: 500; }
.col-author  { max-width: 90px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
.col-summary { max-width: 320px; word-wrap: break-word; line-height: 1.4; }
.col-login   { color: #888; font-size: 12px; }

/* ── Ресайз столбцов ── */
.col-resize-handle {
  position: absolute; right: 0; top: 0;
  width: 5px; height: 100%;
  cursor: col-resize;
  user-select: none;
  z-index: 1;
}
.col-resize-handle:hover,
.col-resize-handle.dragging { background: #1a73e8; }
thead th { cursor: grab; }
thead th.th-drag-over { border-left: 2px solid #1a73e8 !important; }
thead th.th-dragging  { opacity: .45; }

/* Channel colors */
.ch-chat { color: #1a73e8; }
.ch-ls   { color: #0f9d58; }
.ch-form { color: #e67e22; }

/* Editable cells */
.editable {
  display: block;
  min-width: 80px; min-height: 20px;
  padding: 3px 5px; border-radius: 4px;
  outline: none; cursor: text;
  white-space: pre-wrap; word-break: break-word; overflow-wrap: break-word;
  transition: background .15s, box-shadow .15s;
}
.editable:empty::before {
  content: attr(data-placeholder);
  color: #bbb; font-style: italic; pointer-events: none;
}
.editable:empty { cursor: pointer; }
.editable:focus { background: #fffbea; box-shadow: 0 0 0 2px #f59e0b55; }
.editable.saving { opacity: .6; }
.editable.saved  { animation: flash 1s ease forwards; }
.comment-wrap { display: flex; flex-direction: column; gap: 1px; }
.comment-author { font-weight: 600; font-size: 11px; color: #888; line-height: 1.2; }
.comment-ro { font-size: 12px; line-height: 1.4; white-space: pre-wrap; word-break: break-word; }
.btn-append-comment { background: none; border: 1px dashed #ccc; border-radius: 4px; color: #888; font-size: 11px; cursor: pointer; padding: 1px 6px; margin-top: 2px; align-self: flex-start; }
.btn-append-comment:hover { border-color: #888; color: #555; }
.comment-append-area { display: flex; flex-direction: column; gap: 4px; margin-top: 4px; }
.comment-append-input { min-height: 28px; font-size: 12px; padding: 3px 6px; border: 1px solid #ccc; border-radius: 4px; outline: none; white-space: pre-wrap; word-break: break-word; }
.comment-append-input:empty::before { content: attr(data-placeholder); color: #bbb; }
.comment-append-input:focus { border-color: #4a90d9; background: #fffbea; }
.comment-append-actions { display: flex; gap: 6px; }
.btn-app-save { background: #4a90d9; color: #fff; border: none; border-radius: 4px; font-size: 11px; cursor: pointer; padding: 2px 10px; }
.btn-app-save:hover { background: #2a70b9; }
.btn-app-cancel { background: none; border: 1px solid #ccc; border-radius: 4px; font-size: 11px; cursor: pointer; padding: 2px 8px; color: #666; }
mark.hl { background: #ffe566; color: inherit; border-radius: 2px; padding: 0 1px; font-style: normal; }

.select-cell {
  cursor: pointer; padding: 3px 7px; border-radius: 5px;
  min-width: 60px; min-height: 20px;
  transition: background .15s;
  display: flex; align-items: center; gap: 4px;
}
.select-cell:hover { background: #f0f7ff; }
.select-cell::after {
  content: '▾'; font-size: 10px; color: #bbb; flex-shrink: 0;
  transition: color .15s;
}
.select-cell:hover::after { color: #1a73e8; }
.select-cell.saved { animation: flash 1s ease forwards; }
/* Cell dropdown overlay */
.cell-dd {
  position: fixed; z-index: 9999;
  background: #fff; border: 1px solid #e0e0e0; border-radius: 10px;
  box-shadow: 0 8px 32px rgba(0,0,0,.14); padding: 4px 0;
  min-width: 140px; max-height: 260px; overflow-y: auto;
}
.cell-dd-opt {
  padding: 8px 16px; font-size: 13px; color: #333;
  cursor: pointer; transition: background .1s;
  display: flex; align-items: center; gap: 8px;
}
.cell-dd-opt:hover { background: #f0f7ff; color: #1a73e8; }
.cell-dd-opt.active {
  background: #e8f0fe; color: #1a73e8; font-weight: 600;
}
.cell-dd-opt.active::before {
  content: '✓'; font-size: 11px; color: #1a73e8;
}

/* Ответственный отдел — своя подсветка выпадашки */
.select-dept select { border-color: #818cf8; background: #eef2ff; }

@keyframes flash {
  0%   { background: #d1fae5; }
  100% { background: transparent; }
}

/* Result badge */
.badge {
  display: inline-block; padding: 2px 8px; border-radius: 10px;
  font-size: 11px; font-weight: 600; white-space: nowrap;
}
.badge-решено     { background: #d1fae5; color: #065f46; }
.badge-не-решено  { background: #fee2e2; color: #991b1b; }
.badge-частично   { background: #fef3c7; color: #92400e; }
.badge-эскалация  { background: #ede9fe; color: #5b21b6; }
.badge-other      { background: #f3f4f6; color: #374151; }

/* Dept badge */
.dept-badge {
  display: inline-block; padding: 2px 8px; border-radius: 10px;
  font-size: 11px; font-weight: 600; white-space: nowrap;
  background: #e0e7ff; color: #3730a3;
}

/* Rating / Complexity widget */
.rating-widget { display: flex; gap: 2px; align-items: center; }
.rating-dot {
  width: 22px; height: 22px; border-radius: 50%;
  display: flex; align-items: center; justify-content: center;
  font-size: 11px; font-weight: 700; cursor: default;
  border: 1.5px solid #e0e0e0; color: #bbb; background: #fafafa;
  transition: background .15s, border-color .15s, color .15s;
  user-select: none; flex-shrink: 0;
}
.rating-dot.filled-rating     { background: #e8f5e9; border-color: #66bb6a; color: #2e7d32; }
.rating-dot.filled-complexity { background: #fff3e0; border-color: #ffa726; color: #e65100; }
.rating-dot.clickable { cursor: pointer; }
.rating-dot.clickable:hover   { border-color: #1a73e8; color: #1a73e8; background: #e8f0fe; }

/* Problem summary truncation */
.summary-text {
  display: -webkit-box;
  -webkit-line-clamp: 3;
  -webkit-box-orient: vertical;
  overflow: hidden; cursor: pointer;
}
.summary-text.expanded {
  display: block; -webkit-line-clamp: unset;
}

.no-data { text-align: center; padding: 60px 20px; color: #aaa; font-size: 14px; }
.spinner { display: inline-block; width: 16px; height: 16px; border: 2px solid #ddd; border-top-color: #1a73e8; border-radius: 50%; animation: spin .6s linear infinite; vertical-align: middle; margin-right: 6px; }
@keyframes spin { to { transform: rotate(360deg); } }

/* ── Пагинация ── */
.pagination {
  flex-shrink: 0;
  display: flex; align-items: center; justify-content: center; gap: 4px;
  padding: 8px; background: #fff; border-top: 1px solid #e8e8e8;
}
.pg-btn {
  min-width: 32px; height: 28px; padding: 0 8px;
  border: 1px solid #ddd; border-radius: 4px; background: #fff;
  font-size: 12px; cursor: pointer; color: #333;
  display: flex; align-items: center; justify-content: center;
}
.pg-btn:hover:not(:disabled) { background: #f0f7ff; border-color: #1a73e8; color: #1a73e8; }
.pg-btn.active { background: #1a73e8; color: #fff; border-color: #1a73e8; font-weight: 600; }
.pg-btn:disabled { opacity: .4; cursor: default; }
.pg-info { font-size: 12px; color: #888; padding: 0 8px; }
</style>
</head>
<body>

<div class="topbar">
  <h1>📊 Day Tracker</h1>
  <div class="topbar-right">
    <div class="topbar-search">
      <span class="topbar-search-icon">🔍</span>
      <input type="text" id="filter_search" placeholder="Поиск по всем полям...">
    </div>
    %%ADMIN_LINK%%
    <div class="user-chip">
      <div class="user-avatar" id="user-avatar"></div>
      <span>%%USER%%</span>
    </div>
    <a href="/auth/logout" class="topbar-link">Выйти</a>
  </div>
</div>
<div class="filterbar">
  <div class="fg">
    <label>С</label>
    <input type="date" id="date_from">
  </div>
  <div class="fg">
    <label>По</label>
    <input type="date" id="date_to">
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Оператор</label>
    <select id="filter_operator"><option value="">Все</option></select>
  </div>
  <div class="fg">
    <label>Автор</label>
    <input type="text" id="filter_author" placeholder="введите имя" style="width:130px">
  </div>
  <div class="fg">
    <label>Логин</label>
    <input type="text" id="filter_login" placeholder="введите логин" style="width:130px">
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Тип</label>
    <select id="filter_stype">
      <option value="">Все</option>
      <option value="Клиент">Клиент</option>
      <option value="ПВЗ">ПВЗ</option>
      <option value="Сотрудник">Сотрудник</option>
    </select>
  </div>
  <div class="fg">
    <label>Канал</label>
    <select id="filter_channel"><option value="">Все</option></select>
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Категория</label>
    <select id="filter_category" onchange="updateSubcatFilter()">
      <option value="">Все</option>
    </select>
  </div>
  <div class="fg">
    <label>Подкатегория</label>
    <select id="filter_subcategory"><option value="">Все</option></select>
  </div>
  <div class="filter-sep"></div>
  <div class="fg">
    <label>Результат</label>
    <select id="filter_result">
      <option value="">Все</option>
      <option value="Решено">Решено</option>
      <option value="Не решено">Не решено</option>
      <option value="Эскалация">Эскалация</option>
    </select>
  </div>
  <div class="fg">
    <label>Отдел</label>
    <select id="filter_dept">
      <option value="">Все</option>
    </select>
  </div>
  <div class="filterbar-actions">
    <button class="btn" onclick="load()">Применить</button>
    <span class="filterbar-count" id="count"></span>
  </div>
</div>

<div class="wrap">
  <table>
    <thead>
      <tr>
        <th data-col="date">Дата</th>
        <th data-col="time">Время</th>
        <th data-col="operator">Оператор</th>
        <th data-col="source_type">Тип автора</th>
        <th data-col="author">Автор</th>
        <th data-col="login">Логин</th>
        <th data-col="appeal_type">Тип</th>
        <th data-col="category">Категория</th>
        <th data-col="subcategory">Подкатегория</th>
        <th data-col="problem_summary">Причина обращения</th>
        <th data-col="result">Результат</th>
        <th data-col="comment">Комм. поддержки</th>
        <th data-col="comment_manager">Комм. менеджера</th>
        <th data-col="responsible_dept" class="col-dept">Отв. отдел</th>
        <th data-col="channel">Канал</th>
      </tr>
    </thead>
    <tbody id="tbody">
      <tr><td colspan="15" class="no-data"><span class="spinner"></span>Загрузка...</td></tr>
    </tbody>
  </table>
</div>

<div class="pagination" id="pagination"></div>

<script>
const CURRENT_USER = "%%USER%%";
const PERMS        = %%PERMS%%;

// Аватар — первая буква имени
(function() {
  const av = document.getElementById('user-avatar');
  if (av && CURRENT_USER) av.textContent = CURRENT_USER[0].toUpperCase();
})();

const today = new Date();
const week  = new Date(today); week.setDate(today.getDate() - 7);
document.getElementById('date_from').value = fmt(week);
document.getElementById('date_to').value   = fmt(today);

let currentPage = 1;

function fmt(d) { return d.toISOString().split('T')[0]; }

function fmtDate(s) {
  if (!s) return '';
  const p = s.split('-');
  if (p.length === 3) return p[2] + '.' + p[1] + '.' + p[0];
  return s;
}

function esc(s) {
  if (s === null || s === undefined) return '';
  return String(s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;')
    .replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

function badgeClass(result) {
  const r = (result || '').toLowerCase();
  if (r === 'решено')    return 'badge-решено';
  if (r === 'не решено') return 'badge-не-решено';
  if (r === 'частично')  return 'badge-частично';
  if (r === 'эскалация') return 'badge-эскалация';
  return r ? 'badge-other' : '';
}

// ── Константы категорий ────────────────────────────────────────────────────
const SOURCE_TYPES = ['Клиент', 'ПВЗ', 'Сотрудник'];

const DEPTS = [
  'Логистика',
  'Финансы',
  'Технический отдел',
  'Клиентский сервис',
  'Операции / ПВЗ',
  'Контент',
  'Маркетинг',
];

const CAT_MAP = {
  'Заказ и доставка':      ['Статус заказа / задержка','Заказ не найден / потерян','Статус не обновился','Объединение / формирование посылки','Возврат товара (логистика)'],
  'ПВЗ и выдача':          ['График / адрес ПВЗ','Заказ не выдается','Проблема при приеме товара','Не сканируется / нет стикера','Ручные операции / обход процесса','Стать раздающим'],
  'Оплата и финансы':      ['Возврат денег','Непонятные списания / долг','Платное хранение','Оргсбор — размер и расчёт'],
  'Технические проблемы':  ['Ошибка / сбой сайта','Некорректное отображение','Поиск / каталог','Мобильное приложение'],
  'Аккаунт и профиль':     ['Нет доступа / ошибка входа','Смена телефона / почты','Блокировка / ограничения','Управление рассылками'],
  'Качество товара':       ['Брак / дефект','Несоответствие описанию или фото','Пересорт / пришло не то','Повреждение при доставке','Нарушения — маркировка, сроки годности'],
  'Закупки и организаторы':['Вопрос по закупке','Статус закупки / оплаты','Жалоба на организатора'],
  'Пристрой':              ['Как работает пристрой','Проблемы / передача пристроя'],
  'Пожелания и инсайты':   ['Запрос нового бренда / поставщика','Запрос открытия ПВЗ в регионе','Запрос новой функции платформы','Коммерческое предложение / партнёрство','Сравнение с конкурентами (WB, Ozon)'],
  'Благодарность':         ['Благодарность'],
  'Не определено':         ['Служебное / внутреннее'],
};

function initFilters() {
  const sel = document.getElementById('filter_category');
  sel.innerHTML = '<option value="">Все</option>' +
    Object.keys(CAT_MAP).map(c => `<option value="${esc(c)}">${esc(c)}</option>`).join('');

  const dsel = document.getElementById('filter_dept');
  dsel.innerHTML = '<option value="">Все</option>' +
    DEPTS.map(d => `<option value="${esc(d)}">${esc(d)}</option>`).join('');
}

function updateSubcatFilter() {
  const cat = document.getElementById('filter_category').value;
  const sel = document.getElementById('filter_subcategory');
  const cur = sel.value;
  const subs = CAT_MAP[cat] || [];
  sel.innerHTML = '<option value="">Все</option>' +
    subs.map(s => `<option value="${esc(s)}"${s===cur?' selected':''}>${esc(s)}</option>`).join('');
}

function goPage(p) { currentPage = p; load(false); }

async function load(resetPage = true) {
  if (resetPage) currentPage = 1;

  const tbody = document.getElementById('tbody');
  tbody.innerHTML = '<tr><td colspan="15" class="no-data"><span class="spinner"></span>Загрузка...</td></tr>';
  document.getElementById('count').textContent = '';
  document.getElementById('pagination').innerHTML = '';

  try {
    const dr = id => { const el = document.getElementById(id); return el ? (el.value || el.dataset.restore || '') : ''; };
    const gv = id => { const el = document.getElementById(id); return el ? el.value : ''; };

    const params = new URLSearchParams({
      date_from:   gv('date_from'),
      date_to:     gv('date_to'),
      operator:    dr('filter_operator'),
      channel:     dr('filter_channel'),
      stype:       gv('filter_stype'),
      category:    gv('filter_category'),
      subcategory: gv('filter_subcategory'),
      result:      gv('filter_result'),
      dept:        gv('filter_dept'),
      author:      gv('filter_author'),
      login:       gv('filter_login'),
      search:      gv('filter_search'),
      page:        currentPage,
    });

    const resp = await fetch('/api/day-tracker?' + params);
    if (!resp.ok) throw new Error('HTTP ' + resp.status);
    const data = await resp.json();

    if (resetPage) {
      const selOp = document.getElementById('filter_operator');
      const curOp = selOp.value;
      selOp.innerHTML = '<option value="">Все</option>' +
        data.operators.map(o => `<option value="${esc(o)}"${o===curOp?' selected':''}>${esc(o)}</option>`).join('');

      const selCh = document.getElementById('filter_channel');
      const curCh = selCh.value;
      selCh.innerHTML = '<option value="">Все</option>' +
        (data.channels||[]).map(c => `<option value="${esc(c)}"${c===curCh?' selected':''}>${esc(c)}</option>`).join('');
    }

    render(data.rows);
    highlightSearch((document.getElementById('filter_search') || {}).value || '');
    applyHiddenCols();
    renderPagination(data.page, data.pages, data.total);
    document.getElementById('count').textContent =
      `${data.total} записей`;
  } catch (e) {
    tbody.innerHTML = `<tr><td colspan="15" class="no-data">Ошибка загрузки: ${e.message}</td></tr>`;
  }
}

function highlightSearch(term) {
  if (!term || term.length < 3) return;
  const tbody = document.getElementById('tbody');
  const lower = term.toLowerCase();

  function walkNode(node) {
    if (node.nodeType === 3) {
      const text = node.textContent;
      const ltext = text.toLowerCase();
      if (ltext.indexOf(lower) === -1) return;
      const frag = document.createDocumentFragment();
      let pos = 0, i;
      while ((i = ltext.indexOf(lower, pos)) !== -1) {
        if (i > pos) frag.appendChild(document.createTextNode(text.slice(pos, i)));
        const mark = document.createElement('mark');
        mark.className = 'hl';
        mark.textContent = text.slice(i, i + term.length);
        frag.appendChild(mark);
        pos = i + term.length;
      }
      if (pos < text.length) frag.appendChild(document.createTextNode(text.slice(pos)));
      node.parentNode.replaceChild(frag, node);
    } else if (node.nodeType === 1) {
      if (node.contentEditable === 'true' || node.tagName === 'SELECT' || node.tagName === 'INPUT') return;
      Array.from(node.childNodes).forEach(walkNode);
    }
  }

  tbody.querySelectorAll('td').forEach(td => walkNode(td));
}

function applyHiddenCols() {
  const hidden = (PERMS.hidden || []);
  if (!hidden.length) return;
  const allThs = [...document.querySelectorAll('thead th')];
  hidden.forEach(function(col) {
    const th = document.querySelector('thead th[data-col="' + col + '"]');
    if (!th) return;
    const idx = allThs.indexOf(th) + 1;
    const s = document.createElement('style');
    s.textContent = 'thead th:nth-child(' + idx + '), tbody td:nth-child(' + idx + ') { display:none !important; }';
    document.head.appendChild(s);
  });
}

function renderPagination(page, pages, total) {
  const el = document.getElementById('pagination');
  if (pages <= 1) { el.innerHTML = ''; return; }
  const btns = [];
  btns.push(`<button class="pg-btn" onclick="goPage(${page-1})" ${page===1?'disabled':''}>‹</button>`);
  const range = new Set([1, pages]);
  for (let i = Math.max(1, page-2); i <= Math.min(pages, page+2); i++) range.add(i);
  let prev = 0;
  for (const p of [...range].sort((a,b)=>a-b)) {
    if (prev && p - prev > 1) btns.push(`<span class="pg-info">…</span>`);
    btns.push(`<button class="pg-btn${p===page?' active':''}" onclick="goPage(${p})">${p}</button>`);
    prev = p;
  }
  btns.push(`<button class="pg-btn" onclick="goPage(${page+1})" ${page===pages?'disabled':''}>›</button>`);
  el.innerHTML = btns.join('');
}

function commentCell(field, value, author, placeholder) {
  const authorHtml = author ? `<span class="comment-author">${esc(author)}:</span>` : '';
  const isOwner = !author || author === CURRENT_USER;
  const isLead  = PERMS && PERMS.is_lead;
  if (!value || isOwner || isLead) {
    return `<div class="comment-wrap">${authorHtml}<div class="editable" contenteditable="true" data-field="${field}" data-orig="${esc(value)}" data-orig-author="${esc(author)}" data-placeholder="${placeholder}">${esc(value)}</div></div>`;
  }
  return `<div class="comment-wrap">${authorHtml}` +
    `<div class="comment-ro">${esc(value)}</div>` +
    `<span class="editable" data-field="${field}" data-orig="${esc(value)}" data-orig-author="${esc(author)}" style="display:none">${esc(value)}</span>` +
    `<button class="btn-append-comment" onclick="showAppend(this)">+ Добавить</button>` +
    `<div class="comment-append-area" style="display:none">` +
      `<div class="comment-append-input" contenteditable="true" data-placeholder="Ваш комментарий..."></div>` +
      `<div class="comment-append-actions">` +
        `<button class="btn-app-save" onclick="saveAppend(this,'${field}')">Сохранить</button>` +
        `<button class="btn-app-cancel" onclick="hideAppend(this)">Отмена</button>` +
      `</div>` +
    `</div>` +
  `</div>`;
}

function showAppend(btn) {
  btn.style.display = 'none';
  const area = btn.nextElementSibling;
  area.style.display = '';
  area.querySelector('.comment-append-input').focus();
}

function hideAppend(anyBtn) {
  const area    = anyBtn.closest('.comment-append-area');
  const showBtn = area.previousElementSibling;
  area.querySelector('.comment-append-input').textContent = '';
  area.style.display = 'none';
  showBtn.style.display = '';
}

async function saveAppend(btn, field) {
  const area    = btn.closest('.comment-append-area');
  const input   = area.querySelector('.comment-append-input');
  const newText = input.textContent.trim();
  if (!newText) { hideAppend(btn); return; }
  const wrap   = area.closest('.comment-wrap');
  const row    = area.closest('tr');
  const chatId = row.dataset.id;
  const hiddenSpan   = wrap.querySelector(`.editable[data-field="${field}"]`);
  const existingText = hiddenSpan ? hiddenSpan.textContent : '';
  const combined     = existingText + '\\n\\n' + CURRENT_USER + ': ' + newText;
  const allFields    = collectFields(row);
  allFields[field]   = combined;
  const authorSpan = wrap.querySelector('.comment-author');
  const origAuthor = authorSpan ? authorSpan.textContent.replace(/:$/, '').trim() : '';
  if (field === 'comment')         allFields.comment_author         = origAuthor;
  if (field === 'comment_manager') allFields.comment_manager_author = origAuthor;
  const otherField     = field === 'comment' ? 'comment_manager' : 'comment';
  const otherAuthorKey = field === 'comment' ? 'comment_manager_author' : 'comment_author';
  const otherEl        = row.querySelector(`.editable[data-field="${otherField}"]`);
  if (otherEl) {
    const otherWrap = otherEl.closest && otherEl.closest('.comment-wrap');
    const otherSp   = otherWrap && otherWrap.querySelector('.comment-author');
    allFields[otherAuthorKey] = otherSp ? otherSp.textContent.replace(/:$/, '').trim() : '';
  }
  btn.disabled = true;
  try {
    const r = await fetch('/api/day-tracker/' + chatId, { method: 'POST', headers: {'Content-Type': 'application/json'}, body: JSON.stringify(allFields) });
    if (r.ok) {
      if (hiddenSpan) { hiddenSpan.textContent = combined; hiddenSpan.dataset.orig = combined; }
      const roDiv = wrap.querySelector('.comment-ro');
      if (roDiv) roDiv.textContent = combined;
      hideAppend(btn);
    }
  } catch(e) { console.error(e); }
  btn.disabled = false;
}

function render(rows) {
  const tbody = document.getElementById('tbody');
  if (!rows.length) {
    tbody.innerHTML = '<tr><td colspan="15" class="no-data">Нет данных за выбранный период</td></tr>';
    return;
  }

  tbody.innerHTML = rows.map(r => {
    const chClass  = r.channel === 'Чат' ? 'ch-chat' : 'ch-ls';
    const bc       = badgeClass(r.result);
    const resultHtml = bc
      ? `<span class="badge ${bc}">${esc(r.result)}</span>`
      : (r.result ? esc(r.result) : '<span style="color:#bbb">—</span>');
    const deptHtml = r.responsible_dept
      ? `<span class="dept-badge">${esc(r.responsible_dept)}</span>`
      : '<span style="color:#bbb">—</span>';
    return `<tr data-id="${r.chat_id}">
      <td class="col-date">${fmtDate(r.date)}</td>
      <td class="col-time">${esc(r.time)}</td>
      <td>${esc(r.operator)}</td>
      <td><div class="select-cell" data-field="source_type" data-value="${esc(r.source_type)}" onclick="openSelect(this)">${esc(r.source_type) || '<span style=color:#bbb>—</span>'}</div></td>
      <td class="col-author" title="${esc(r.author)}">${esc(r.author)}</td>
      <td class="col-login">${esc(extractLogin(r.author, r.source_type))}</td>
      <td>${esc(r.appeal_type)}</td>
      <td><div class="select-cell" data-field="category"    data-value="${esc(r.category)}"    onclick="openSelect(this)">${esc(r.category)    || '<span style=color:#bbb>—</span>'}</div></td>
      <td><div class="select-cell" data-field="subcategory" data-value="${esc(r.subcategory)}" onclick="openSelect(this)">${esc(r.subcategory) || '<span style=color:#bbb>—</span>'}</div></td>
      <td class="col-summary"><div class="summary-text" onclick="this.classList.toggle('expanded')">${esc(r.problem_summary)}</div></td>
      <td>${resultHtml}</td>
      <td>${commentCell('comment', r.comment, r.comment_author, 'Добавить...')}</td>
      <td>${commentCell('comment_manager', r.comment_manager||'', r.comment_manager_author||'', 'Добавить...')}</td>
      <td class="col-dept"><div class="select-cell select-dept" data-field="responsible_dept" data-value="${esc(r.responsible_dept)}" onclick="openSelect(this)">${deptHtml}</div></td>
      <td class="col-channel ${chClass}">${esc(r.channel)}</td>
    </tr>`;
  }).join('');

  tbody.querySelectorAll('.editable').forEach(el => {
    el.addEventListener('blur', onBlur);
  });
}

async function onBlur(e) {
  const el   = e.target;
  const orig = el.dataset.orig;
  const val  = el.textContent.trim();
  if (val === orig) return;
  if (val === '') el.innerHTML = '';
  el.dataset.orig = val;
  el.classList.add('saving');
  await saveRow(el.closest('tr'), el);
  el.classList.remove('saving');

  const field = el.dataset.field;
  if (field === 'comment' || field === 'comment_manager') {
    const wrap = el.closest('.comment-wrap');
    if (wrap) {
      let authorSpan = wrap.querySelector('.comment-author');
      if (val) {
        if (!authorSpan) {
          authorSpan = document.createElement('span');
          authorSpan.className = 'comment-author';
          wrap.insertBefore(authorSpan, wrap.firstChild);
        }
        const origAuthor = el.dataset.origAuthor || '';
        authorSpan.textContent = (origAuthor || CURRENT_USER) + ':';
      } else if (authorSpan) {
        authorSpan.remove();
      }
    }
  }
}

function extractLogin(author, sourceType) {
  if (!author) return '';
  if (sourceType === 'ПВЗ' || sourceType === 'Сотрудник') return author;
  const ci = author.indexOf(',');
  if (ci !== -1) return author.slice(ci + 1).trim();
  const li = author.toLowerCase().indexOf('лог');
  if (li !== -1) {
    const after = author.slice(li + 3).replace(/^[\\s:]+/, '').trim();
    if (after) return after;
  }
  const words = author.trim().split(/\\s+/);
  if (words.length) return words[words.length - 1];
  return author;
}

function collectFields(row) {
  const f = {};
  row.querySelectorAll('.editable').forEach(c => { f[c.dataset.field] = c.textContent.trim(); });
  row.querySelectorAll('.select-cell[data-field]').forEach(c => { f[c.dataset.field] = c.dataset.value || ''; });
  return f;
}

async function saveRow(row, feedbackEl) {
  const chatId = row.dataset.id;
  const fields = collectFields(row);
  const commentEl    = row.querySelector('[data-field="comment"].editable');
  const commentMgrEl = row.querySelector('[data-field="comment_manager"].editable');
  function resolveAuthor(el, newVal, authorKey) {
    if (!el) return;
    if (newVal !== el.dataset.orig) {
      const origAuthor = el.dataset.origAuthor || '';
      if (PERMS.is_lead && origAuthor && origAuthor !== CURRENT_USER) {
        fields[authorKey] = origAuthor + ' (ред. ' + CURRENT_USER + ')';
      } else {
        fields[authorKey] = newVal ? CURRENT_USER : '';
      }
      el.dataset.origAuthor = fields[authorKey];
    } else {
      const sp = el.closest('.comment-wrap') && el.closest('.comment-wrap').querySelector('.comment-author');
      fields[authorKey] = sp ? sp.textContent.replace(/:$/, '').trim() : '';
    }
  }
  resolveAuthor(commentEl,    fields.comment,         'comment_author');
  resolveAuthor(commentMgrEl, fields.comment_manager, 'comment_manager_author');
  try {
    const resp = await fetch('/api/day-tracker/' + chatId, {
      method:  'POST',
      headers: { 'Content-Type': 'application/json' },
      body:    JSON.stringify(fields),
    });
    if (resp.ok && feedbackEl) {
      feedbackEl.classList.add('saved');
      setTimeout(() => feedbackEl.classList.remove('saved'), 1200);
    }
  } catch (err) {
    console.error('Save error:', err);
  }
}

let _activeCellDd = null;

function _cellDdOutside(e) {
  if (!_activeCellDd) return;
  if (_activeCellDd.contains(e.target)) return;
  if (e.target.closest && e.target.closest('.select-cell')) return;
  closeCellDd();
}

function closeCellDd() {
  if (_activeCellDd) {
    _activeCellDd.remove();
    _activeCellDd = null;
    document.removeEventListener('click', _cellDdOutside);
  }
}

function openSelect(cell) {
  closeCellDd();

  const field   = cell.dataset.field;
  const current = cell.dataset.value || '';
  const row     = cell.closest('tr');

  let options = [];
  if (field === 'source_type') {
    options = SOURCE_TYPES;
  } else if (field === 'category') {
    options = Object.keys(CAT_MAP);
  } else if (field === 'subcategory') {
    const catCell = row.querySelector('[data-field="category"]');
    options = CAT_MAP[catCell ? catCell.dataset.value : ''] || [];
  } else if (field === 'responsible_dept') {
    options = DEPTS;
  }

  if (!options.length) return;

  const rect = cell.getBoundingClientRect();
  const dd = document.createElement('div');
  dd.className = 'cell-dd';
  dd.style.top  = (rect.bottom + window.scrollY + 2) + 'px';
  dd.style.left = (rect.left + window.scrollX) + 'px';

  options.forEach(v => {
    const opt = document.createElement('div');
    opt.className = 'cell-dd-opt' + (v === current ? ' active' : '');
    opt.textContent = v;
    opt.addEventListener('mousedown', async e => {
      e.preventDefault();
      e.stopPropagation();
      closeCellDd();
      cell.dataset.value = v;

      if (field === 'responsible_dept') {
        cell.innerHTML = `<span class="dept-badge">${esc(v)}</span>`;
      } else {
        cell.innerHTML = esc(v);
      }

      if (field === 'category') {
        const subCell = row.querySelector('[data-field="subcategory"]');
        if (subCell) {
          subCell.dataset.value = '';
          subCell.innerHTML = '<span style="color:#bbb">—</span>';
        }
      }

      await saveRow(row, cell);
    });
    dd.appendChild(opt);
  });

  document.body.appendChild(dd);
  _activeCellDd = dd;

  setTimeout(() => {
    document.addEventListener('click', _cellDdOutside);
  }, 0);
}

// ── Ресайз столбцов ────────────────────────────────────────────────────────
const COL_WIDTHS_KEY = 'day_tracker_col_widths_v1';
// Дата, Время, Оператор, Тип автора, Автор, Логин, Тип, Категория, Подкатегория, Причина, Результат, Отв.отдел, Комментарий, Канал
const DEFAULT_COL_WIDTHS = [90, 55, 90, 90, 100, 110, 90, 120, 140, 340, 90, 120, 140, 140, 55];

function initResizableColumns() {
  const ths = [...document.querySelectorAll('thead th')];
  ths.forEach((th, i) => {
    // ── Ресайз ──
    const handle = document.createElement('div');
    handle.className = 'col-resize-handle';
    th.appendChild(handle);
    let startX, startW;
    handle.addEventListener('mousedown', e => {
      startX = e.pageX; startW = th.offsetWidth;
      handle.classList.add('dragging');
      document.body.style.cursor = 'col-resize';
      document.body.style.userSelect = 'none';
      const onMove = e => {
        const w = Math.max(30, startW + e.pageX - startX);
        th.style.width = th.style.minWidth = th.style.maxWidth = w + 'px';
      };
      const onUp = () => {
        handle.classList.remove('dragging');
        document.body.style.cursor = '';
        document.body.style.userSelect = '';
        document.removeEventListener('mousemove', onMove);
        document.removeEventListener('mouseup', onUp);
        saveColWidths();
      };
      document.addEventListener('mousemove', onMove);
      document.addEventListener('mouseup', onUp);
      e.preventDefault();
    });

    // ── Drag-to-reorder ──
    if (!th.dataset.col) return;
    th.draggable = true;
    th.addEventListener('dragstart', e => {
      e.dataTransfer.setData('text/plain', th.dataset.col);
      e.dataTransfer.effectAllowed = 'move';
      setTimeout(() => th.classList.add('th-dragging'), 0);
    });
    th.addEventListener('dragend', () => {
      th.classList.remove('th-dragging');
      document.querySelectorAll('thead th').forEach(t => t.classList.remove('th-drag-over'));
    });
    th.addEventListener('dragover', e => {
      if (!e.dataTransfer.types.includes('text/plain')) return;
      e.preventDefault();
      e.dataTransfer.dropEffect = 'move';
      document.querySelectorAll('thead th').forEach(t => t.classList.remove('th-drag-over'));
      if (th.dataset.col !== e.dataTransfer.getData('text/plain')) th.classList.add('th-drag-over');
    });
    th.addEventListener('dragleave', () => th.classList.remove('th-drag-over'));
    th.addEventListener('drop', e => {
      e.preventDefault();
      th.classList.remove('th-drag-over');
      const fromCol = e.dataTransfer.getData('text/plain');
      const toCol   = th.dataset.col;
      if (!fromCol || fromCol === toCol) return;
      const ord = loadColOrder();
      const fi  = ord.indexOf(fromCol);
      const ti  = ord.indexOf(toCol);
      if (fi === -1 || ti === -1) return;
      ord.splice(fi, 1);
      ord.splice(ti, 0, fromCol);
      saveColOrder(ord);
      applyColLayout();
    });
  });
  restoreColWidths();
}

function saveColWidths() {
  const widths = [...document.querySelectorAll('thead th')].map(th => th.style.width || '');
  localStorage.setItem(COL_WIDTHS_KEY, JSON.stringify(widths));
}

function restoreColWidths() {
  try {
    const saved = JSON.parse(localStorage.getItem(COL_WIDTHS_KEY) || '[]');
    const ths = [...document.querySelectorAll('thead th')];
    ths.forEach((th, i) => {
      const w = saved[i] || (DEFAULT_COL_WIDTHS[i] ? DEFAULT_COL_WIDTHS[i] + 'px' : null);
      if (w) th.style.width = th.style.minWidth = th.style.maxWidth = w;
    });
  } catch (_) {}
}

try { initFilters(); } catch(e) { console.error('initFilters error:', e); }
try { initResizableColumns(); } catch(e) { console.error('initResizableColumns error:', e); }

// Автопоиск по Автору, Логину и Поиску от 3 символов с debounce 400ms
let _searchTimer = null;
['filter_author', 'filter_login', 'filter_search'].forEach(id => {
  const el = document.getElementById(id);
  if (!el) return;
  el.addEventListener('input', function() {
    clearTimeout(_searchTimer);
    const val = this.value;
    if (val.length === 0 || val.length >= 3) {
      _searchTimer = setTimeout(() => load(), 400);
    }
  });
});

load();
</script>
</body>
</html>
"""
